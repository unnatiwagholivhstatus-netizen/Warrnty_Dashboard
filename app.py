import pandas as pd
import numpy as np
from datetime import datetime
import uvicorn
from fastapi import FastAPI, Request, HTTPException
from fastapi.responses import HTMLResponse, StreamingResponse
import os
import sys
import io
import traceback
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

IS_RENDER = os.getenv('RENDER', 'false').lower() == 'true'
DATA_DIR = os.getenv('DATA_DIR', '/mnt/data' if IS_RENDER else '.')

print(f"\n{'='*100}")
print(f"WARRANTY MANAGEMENT SYSTEM - FINAL COMPLETE")
print(f"{'='*100}\n")

WARRANTY_DATA = {
    'credit_df': None, 'debit_df': None, 'arbitration_df': None, 'source_df': None,
    'current_month_df': None, 'current_month_source_df': None,
    'compensation_df': None, 'compensation_source_df': None,
    'pr_approval_df': None, 'pr_approval_source_df': None
}

def find_data_file(filename):
    possible_paths = [filename, f"./{filename}", os.path.join(DATA_DIR, filename),
                     os.path.join(DATA_DIR, 'data', filename), os.path.join('data', filename)]
    
    if filename.endswith('.xlsx'):
        name = filename.replace('.xlsx', '')
        copy_variant = f"{name} - Copy.xlsx"
        possible_paths.extend([copy_variant, f"./{copy_variant}", os.path.join(DATA_DIR, copy_variant)])
    
    for path in possible_paths:
        if os.path.exists(path):
            print(f"  [DONE] Found: {filename}")
            return path
    print(f"  [ERROR] Not found: {filename}")
    return None

def style_worksheet(ws, df, header_fill, header_font, border):
    for col_idx, column in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=column)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(value, (int, float)):
                cell.value = value
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right', vertical='center')
            elif isinstance(value, (datetime, pd.Timestamp)):
                cell.value = value
                cell.number_format = 'mm-dd-yyyy'
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.value = str(value) if not pd.isna(value) else ''
                cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = border
    
    for col_idx, column in enumerate(df.columns, 1):
        max_length = min(max(df[column].astype(str).map(len).max(), len(str(column))) + 2, 40)
        ws.column_dimensions[get_column_letter(col_idx)].width = max_length

# ================== DATA PROCESSING ==================

def process_warranty_data():
    input_path = find_data_file('Warranty Debit.xlsx')
    if input_path is None:
        return None, None, None, None
    
    try:
        df = pd.read_excel(input_path, sheet_name='Sheet1')
        print(f"  [DONE] Warranty data loaded: {len(df)} rows")

        dealer_mapping = {
            'AMRAVATI': 'AMT', 'CHAUFULA_SZZ': 'CHA', 'CHIKHALI': 'CHI',
            'KOLHAPUR_WS': 'KOL', 'NAGPUR_KAMPTHEE ROAD': 'HO',
            'NAGPUR_WARDHAMAN NGR': 'CITY', 'SHIKRAPUR_SZS': 'SHI',
            'WAGHOLI': 'WAG', 'YAVATMAL': 'YAT', 'NAGPUR_WARDHAMAN NGR_CQ': 'CQ'
        }

        numeric_columns = ['Total Claim Amount', 'Credit Note Amount', 'Debit Note Amount']
        for col in numeric_columns:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

        df['Dealer_Code'] = df['Dealer Location'].map(dealer_mapping).fillna(df['Dealer Location'])
        df['Month'] = df['Fiscal Month'].astype(str).str.strip().str[:3]
        df['Claim arbitration ID'] = df['Claim arbitration ID'].astype(str).replace('nan', '').replace('', np.nan)

        dealers = sorted(df['Dealer_Code'].unique())
        months = ['Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec', 'Jan']

        # CREDIT
        credit_df = pd.DataFrame({'Division': dealers})
        for month in months:
            month_data = df[df['Month'] == month]
            if not month_data.empty:
                summary = month_data.groupby('Dealer_Code')['Credit Note Amount'].sum().reset_index()
                summary.columns = ['Division', f'Credit Note {month}']
                credit_df = credit_df.merge(summary, on='Division', how='left')
            else:
                credit_df[f'Credit Note {month}'] = 0
        
        credit_df = credit_df.fillna(0)
        credit_columns = [f'Credit Note {month}' for month in months]
        credit_df['Total Credit'] = credit_df[credit_columns].sum(axis=1)
        grand_total_credit = {'Division': 'Grand Total'}
        for col in credit_df.columns[1:]:
            grand_total_credit[col] = credit_df[col].sum()
        credit_df = pd.concat([credit_df, pd.DataFrame([grand_total_credit])], ignore_index=True)

        # DEBIT
        debit_df = pd.DataFrame({'Division': dealers})
        for month in months:
            month_data = df[df['Month'] == month]
            if not month_data.empty:
                summary = month_data.groupby('Dealer_Code')['Debit Note Amount'].sum().reset_index()
                summary.columns = ['Division', f'Debit Note {month}']
                debit_df = debit_df.merge(summary, on='Division', how='left')
            else:
                debit_df[f'Debit Note {month}'] = 0
        
        debit_df = debit_df.fillna(0)
        debit_columns = [f'Debit Note {month}' for month in months]
        debit_df['Total Debit'] = debit_df[debit_columns].sum(axis=1)
        grand_total_debit = {'Division': 'Grand Total'}
        for col in debit_df.columns[1:]:
            grand_total_debit[col] = debit_df[col].sum()
        debit_df = pd.concat([debit_df, pd.DataFrame([grand_total_debit])], ignore_index=True)

        # ARBITRATION - TOTAL ONLY (NOT MONTH-WISE)
        def is_arbitration(value):
            if pd.isna(value): return False
            value = str(value).strip().upper()
            return value.startswith('ARB') and value != 'NAN' and value != ''

        arbitration_data = []
        for dealer in dealers:
            dealer_data = df[df['Dealer_Code'] == dealer]
            total_debit = dealer_data['Debit Note Amount'].sum()
            arb_claimed = dealer_data[dealer_data['Claim arbitration ID'].apply(is_arbitration)]['Debit Note Amount'].sum()
            arb_pending = max(0, total_debit - arb_claimed)
            
            arbitration_data.append({
                'Division': dealer,
                'Total Debit': total_debit,
                'Arbitration Claimed': arb_claimed,
                'Arbitration Not Claimed': arb_pending
            })
        
        arbitration_df = pd.DataFrame(arbitration_data)
        grand_total_arb = {
            'Division': 'Grand Total',
            'Total Debit': arbitration_df['Total Debit'].sum(),
            'Arbitration Claimed': arbitration_df['Arbitration Claimed'].sum(),
            'Arbitration Not Claimed': arbitration_df['Arbitration Not Claimed'].sum()
        }
        arbitration_df = pd.concat([arbitration_df, pd.DataFrame([grand_total_arb])], ignore_index=True)

        print("  [DONE] Warranty processing completed")
        return credit_df, debit_df, arbitration_df, df

    except Exception as e:
        print(f"   Error: {e}")
        return None, None, None, None

def process_current_month_warranty():
    input_path = find_data_file('Pending Warranty Claim Details.xlsx')
    if input_path is None:
        return None, None
    
    try:
        df = pd.read_excel(input_path, sheet_name='Pending Warranty Claim Details')
        print(f"   Current Month loaded: {len(df)} rows")
        
        df['Division'] = df['Division'].astype(str).str.strip()
        df = df[df['Division'].notna() & (df['Division'] != '') & (df['Division'] != 'nan')]

        summary_data = []
        for division in sorted(df['Division'].unique()):
            div_data = df[df['Division'] == division]
            summary_data.append({
                'Division': division,
                'Pending Spares Count': int(div_data['Pending Claims Spares'].notna().sum()),
                'Pending Labour Count': int(div_data['Pending Claims Labour'].notna().sum()),
                'Total Pending Claims': int(div_data['Pending Claims Spares'].notna().sum() + div_data['Pending Claims Labour'].notna().sum())
            })
        
        summary_df = pd.DataFrame(summary_data)
        grand_total = {
            'Division': 'Grand Total',
            'Pending Spares Count': int(summary_df['Pending Spares Count'].sum()),
            'Pending Labour Count': int(summary_df['Pending Labour Count'].sum()),
            'Total Pending Claims': int(summary_df['Total Pending Claims'].sum())
        }
        summary_df = pd.concat([summary_df, pd.DataFrame([grand_total])], ignore_index=True)

        print("   Current Month processing completed")
        return summary_df, df

    except Exception as e:
        print(f"   Error: {e}")
        return None, None

def process_compensation_claim():
    input_path = find_data_file('Transit_Claims_Merged.xlsx')
    if input_path is None:
        return None, None
    
    try:
        df = pd.read_excel(input_path)
        print(f"   Compensation loaded: {len(df)} rows")

        required_columns = [
            'Division', 'RO Id.', 'Registration No.', 'RO Date', 'RO Bill Date',
            'Chassis No.', 'Model Group', 'Claim Amount', 'Claim Date',
            'Request No.', 'Request Date', 'Request Status',
            'Claim Approved Amt.', 'No. of Days'
        ]
        
        available_columns = [col for col in required_columns if col in df.columns]
        if not available_columns:
            return None, None

        df_filtered = df[available_columns].copy()
        
        if 'Division' in df_filtered.columns:
            df_filtered['Division'] = df_filtered['Division'].astype(str).str.strip()
            df_filtered = df_filtered[df_filtered['Division'].notna() & 
                                     (df_filtered['Division'] != '') & 
                                     (df_filtered['Division'] != 'nan')]
        
        if 'RO Id.' in df_filtered.columns:
            def format_ro_id(x):
                if pd.isna(x) or str(x).strip() == '':
                    return ''
                try:
                    return f"RO{str(int(float(x)))}"
                except:
                    return str(x).strip()
            df_filtered['RO Id.'] = df_filtered['RO Id.'].apply(format_ro_id)
        
        numeric_cols = ['Claim Amount', 'Claim Approved Amt.', 'No. of Days']
        for col in numeric_cols:
            if col in df_filtered.columns:
                df_filtered[col] = pd.to_numeric(df_filtered[col], errors='coerce').fillna(0)
        
        date_cols = ['RO Date', 'RO Bill Date', 'Claim Date', 'Request Date']
        for col in date_cols:
            if col in df_filtered.columns:
                df_filtered[col] = pd.to_datetime(df_filtered[col], errors='coerce')
        
        summary_data = []
        if 'Division' in df_filtered.columns:
            for division in sorted(df_filtered['Division'].unique()):
                div_data = df_filtered[df_filtered['Division'] == division]
                summary_row = {'Division': division, 'Total Claims': len(div_data)}
                
                if 'Claim Amount' in df_filtered.columns:
                    summary_row['Total Claim Amount'] = div_data['Claim Amount'].sum()
                
                if 'Claim Approved Amt.' in df_filtered.columns:
                    summary_row['Total Approved Amount'] = div_data['Claim Approved Amt.'].sum()
                
                if 'No. of Days' in df_filtered.columns:
                    summary_row['Avg No. of Days'] = div_data['No. of Days'].mean()
                
                summary_data.append(summary_row)
            
            summary_df = pd.DataFrame(summary_data)
            grand_total = {'Division': 'Grand Total'}
            for col in summary_df.columns[1:]:
                if summary_df[col].dtype in ['int64', 'float64']:
                    grand_total[col] = summary_df[col].sum()
            summary_df = pd.concat([summary_df, pd.DataFrame([grand_total])], ignore_index=True)
        else:
            summary_df = pd.DataFrame()

        print("   Compensation processing completed")
        return summary_df, df_filtered

    except Exception as e:
        print(f"   Error: {e}")
        return None, None

def process_pr_approval():
    input_path = find_data_file('Pr_Approval_Claims_Merged.xlsx')
    if input_path is None:
        return None, None
    
    try:
        df = pd.read_excel(input_path)
        print(f"   PR Approval loaded: {len(df)} rows")

        if 'Division' not in df.columns:
            return None, None

        df['Division'] = df['Division'].astype(str).str.strip()
        df = df[df['Division'].notna() & (df['Division'] != '') & (df['Division'] != 'nan')]

        numeric_cols = ['Total Cost of Repair', 'Req. Claim Amt from M&M', 'App. Claim Amt from M&M']
        for col in numeric_cols:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

        summary_data = []
        for division in sorted(df['Division'].unique()):
            div_data = df[df['Division'] == division]
            summary_row = {'Division': division, 'Total Requests': len(div_data)}
            
            if 'Total Cost of Repair' in df.columns:
                summary_row['Total Cost of Repair'] = div_data['Total Cost of Repair'].sum()
            
            if 'Req. Claim Amt from M&M' in df.columns:
                summary_row['Req. Claim Amt from M&M'] = div_data['Req. Claim Amt from M&M'].sum()
            
            if 'App. Claim Amt from M&M' in df.columns:
                summary_row['Total Approved Amount'] = div_data['App. Claim Amt from M&M'].sum()
            
            summary_data.append(summary_row)
        
        summary_df = pd.DataFrame(summary_data)
        grand_total = {'Division': 'Grand Total'}
        for col in summary_df.columns[1:]:
            if summary_df[col].dtype in ['int64', 'float64']:
                grand_total[col] = summary_df[col].sum()
        summary_df = pd.concat([summary_df, pd.DataFrame([grand_total])], ignore_index=True)

        print("   PR Approval processing completed")
        return summary_df, df

    except Exception as e:
        print(f"   Error: {e}")
        return None, None

# ==================== HTML DASHBOARD ====================

DASHBOARD_HTML = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Warranty Dashboard</title>
<style>
* { margin: 0; padding: 0; box-sizing: border-box; }
html, body { height: 100%; }
body { font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background: linear-gradient(135deg, #f5f5f5 0%, #e0e0e0 100%); min-height: 100vh; }
.navbar { background: linear-gradient(135deg, #FF8C00 0%, #FF6B35 100%); color: white; padding: 15px 0; box-shadow: 0 2px 8px rgba(0,0,0,0.15); position: sticky; top: 0; z-index: 100; }
.navbar h1 { max-width: 1400px; margin: 0 auto; font-size: 20px; padding: 0 20px; font-weight: 700; }
.container { max-width: 1400px; margin: 20px auto; padding: 0 15px; }
.dashboard-content { background: white; border-radius: 12px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); padding: 20px; }
.nav-tabs { border-bottom: 2px solid #FF8C00; margin-bottom: 20px; display: flex; flex-wrap: wrap; gap: 5px; }
.nav-tabs button { color: #666; font-weight: 600; border: none; border-bottom: 3px solid transparent; padding: 10px 12px; cursor: pointer; transition: all 0.3s ease; background: none; font-size: 12px; }
.nav-tabs button:hover { color: #FF8C00; }
.nav-tabs button.active { color: #FF8C00; border-bottom-color: #FF8C00; }
.tab-content { display: none; }
.tab-content.active { display: block; }
.export-section { margin: 20px 0; padding: 15px; background: linear-gradient(135deg, #fff8f3 0%, #ffe8d6 100%); border-radius: 8px; border-left: 5px solid #FF8C00; }
.export-section h3 { color: #FF8C00; margin-bottom: 12px; font-size: 14px; font-weight: 700; }
.export-controls { display: flex; gap: 10px; align-items: center; flex-wrap: wrap; background: white; padding: 12px; border-radius: 6px; }
.export-control-group { display: flex; gap: 6px; align-items: center; }
.export-control-group label { font-weight: 600; color: #333; font-size: 12px; min-width: 70px; }
.export-control-group select { padding: 7px 10px; border: 2px solid #FF8C00; border-radius: 4px; background: white; font-size: 11px; min-width: 120px; }
.export-btn { padding: 8px 18px; background: linear-gradient(135deg, #4CAF50 0%, #45a049 100%); color: white; border: none; border-radius: 4px; cursor: pointer; font-weight: 700; font-size: 12px; }
.export-btn:hover { transform: translateY(-2px); box-shadow: 0 5px 15px rgba(76, 175, 80, 0.3); }
.table-wrapper { overflow-x: auto; }
.data-table { width: 100%; border-collapse: collapse; margin-top: 15px; font-size: 11px; }
.data-table thead th { background: linear-gradient(135deg, #FF8C00 0%, #FF6B35 100%); color: white; padding: 10px 8px; text-align: center; font-weight: 600; }
.data-table tbody td { padding: 8px; border-bottom: 1px solid #e0e0e0; text-align: right; }
.data-table tbody td:first-child { text-align: left; font-weight: 600; }
.data-table tbody tr:last-child { background: #fff8f3; font-weight: 700; border-top: 2px solid #FF8C00; color: #FF8C00; }
.table-title { font-size: 14px; font-weight: 700; color: #FF8C00; margin-bottom: 12px; }
.loading-spinner { display: none; text-align: center; padding: 40px; }
.spinner { border: 4px solid rgba(255,140,0,0.2); border-top: 4px solid #FF8C00; border-radius: 50%; width: 40px; height: 40px; animation: spin 1s linear infinite; margin: 0 auto; }
@keyframes spin { 0% { transform: rotate(0deg); } 100% { transform: rotate(360deg); } }
.error-msg { color: #c62828; padding: 12px; background: #ffebee; border-left: 4px solid #c62828; border-radius: 4px; margin: 10px 0; display: none; font-size: 12px; }
.error-msg.show { display: block; }
</style>
</head>
<body>
<nav class="navbar"><h1>Warranty Management Dashboard</h1></nav>
<div class="container">
<div class="dashboard-content">
<div class="loading-spinner" id="loadingSpinner">
<div class="spinner"></div>
<p style="margin-top: 15px; color: #666; font-size: 12px;">Loading warranty data...</p>
</div>

<div id="warrantyTabs" style="display: none;">
<div class="nav-tabs">
<button class="nav-link active" onclick="switchTab(event, 'credit')">Credit</button>
<button class="nav-link" onclick="switchTab(event, 'debit')">Debit</button>
<button class="nav-link" onclick="switchTab(event, 'arbitration')">Arbitration</button>
<button class="nav-link" onclick="switchTab(event, 'currentmonth')">Current Month</button>
<button class="nav-link" onclick="switchTab(event, 'compensation')">Compensation</button>
<button class="nav-link" onclick="switchTab(event, 'pr_approval')">PR Approval</button>
</div>

<div class="export-section">
<h3>Export to Excel</h3>
<div class="export-controls">
<div class="export-control-group">
<label>Division:</label>
<select id="divisionFilter"><option value="">-- Select --</option><option value="All">All</option></select>
</div>
<div class="export-control-group">
<label>Type:</label>
<select id="exportType">
<option value="credit">Credit</option>
<option value="debit">Debit</option>
<option value="arbitration">Arbitration</option>
<option value="currentmonth">Current Month</option>
<option value="compensation">Compensation</option>
<option value="pr_approval">PR Approval</option>
</select>
</div>
<button onclick="exportToExcel()" class="export-btn">Export</button>
</div>
<div class="error-msg" id="exportError"></div>
</div>

<div id="credit" class="tab-content active">
<div class="table-title">Warranty Credit Note</div>
<div class="table-wrapper"><table class="data-table" id="creditTable"><thead></thead><tbody></tbody></table></div>
</div>

<div id="debit" class="tab-content">
<div class="table-title">Warranty Debit Note</div>
<div class="table-wrapper"><table class="data-table" id="debitTable"><thead></thead><tbody></tbody></table></div>
</div>

<div id="arbitration" class="tab-content">
<div class="table-title">Claim Arbitration</div>
<div class="table-wrapper"><table class="data-table" id="arbitrationTable"><thead></thead><tbody></tbody></table></div>
</div>

<div id="currentmonth" class="tab-content">
<div class="table-title">Current Month Pending</div>
<div class="table-wrapper"><table class="data-table" id="currentMonthTable"><thead></thead><tbody></tbody></table></div>
</div>

<div id="compensation" class="tab-content">
<div class="table-title">Compensation Claim</div>
<div class="table-wrapper"><table class="data-table" id="compensationTable"><thead></thead><tbody></tbody></table></div>
</div>

<div id="pr_approval" class="tab-content">
<div class="table-title">PR Approval</div>
<div class="table-wrapper"><table class="data-table" id="prApprovalTable"><thead></thead><tbody></tbody></table></div>
</div>
</div>
</div>
</div>

<script>
let warrantyData = {};

async function loadDashboard() {
    const spinner = document.getElementById('loadingSpinner');
    const tabs = document.getElementById('warrantyTabs');
    
    try {
        const response = await fetch('/api/warranty-data');
        if (!response.ok) throw new Error('Failed');
        
        warrantyData = await response.json();
        
        populateTable(warrantyData.credit, 'creditTable');
        populateTable(warrantyData.debit, 'debitTable');
        populateTable(warrantyData.arbitration, 'arbitrationTable');
        populateTable(warrantyData.currentMonth, 'currentMonthTable');
        populateTable(warrantyData.compensation, 'compensationTable');
        populateTable(warrantyData.prApproval, 'prApprovalTable');
        
        loadDivisions();
        
        spinner.style.display = 'none';
        tabs.style.display = 'block';
    } catch (error) {
        spinner.innerHTML = '<p style="color: red;">Error loading data</p>';
    }
}

function populateTable(data, tableId) {
    if (!data || data.length === 0) return;
    const table = document.getElementById(tableId);
    const headers = Object.keys(data[0]);
    table.querySelector('thead').innerHTML = headers.map(h => `<th>${h}</th>`).join('');
    table.querySelector('tbody').innerHTML = data.map((row) =>
        `<tr>${headers.map((h) => {
            const val = row[h];
            const formatted = typeof val === 'number' ? val.toLocaleString('en-IN', {maximumFractionDigits: 0}) : val;
            return `<td>${formatted}</td>`;
        }).join('')}</tr>`
    ).join('');
}

function switchTab(e, tabName) {
    document.querySelectorAll('.tab-content').forEach(t => t.classList.remove('active'));
    document.querySelectorAll('.nav-link').forEach(b => b.classList.remove('active'));
    document.getElementById(tabName).classList.add('active');
    e.target.classList.add('active');
}

function loadDivisions() {
    const divisions = new Set();
    const type = document.getElementById('exportType').value || 'credit';
    const dataKey = type === 'currentmonth' ? 'currentMonth' : type === 'compensation' ? 'compensation' : type === 'pr_approval' ? 'prApproval' : type;
    const data = warrantyData[dataKey];
    
    if (data) {
        data.forEach(row => {
            if (row.Division && row.Division !== 'Grand Total') {
                divisions.add(row.Division);
            }
        });
    }
    
    const select = document.getElementById('divisionFilter');
    select.innerHTML = '<option value="">-- Select --</option><option value="All">All</option>';
    
    Array.from(divisions).sort().forEach(div => {
        const opt = document.createElement('option');
        opt.value = div;
        opt.textContent = div;
        select.appendChild(opt);
    });
}

document.getElementById('exportType')?.addEventListener('change', loadDivisions);

async function exportToExcel() {
    const division = document.getElementById('divisionFilter').value;
    const type = document.getElementById('exportType').value;
    const errorDiv = document.getElementById('exportError');
    
    errorDiv.classList.remove('show');
    
    if (!division) {
        errorDiv.textContent = '⚠ Select division first';
        errorDiv.classList.add('show');
        return;
    }
    
    try {
        const response = await fetch('/api/export-to-excel', {
            method: 'POST',
            headers: {'Content-Type': 'application/json'},
            body: JSON.stringify({division, type})
        });
        
        if (!response.ok) throw new Error('Export failed');
        
        const blob = await response.blob();
        if (blob.size === 0) throw new Error('Empty file');
        
        const url = URL.createObjectURL(blob);
        const a = document.createElement('a');
        a.href = url;
        a.download = `${type}_${division}_${new Date().toISOString().split('T')[0]}.xlsx`;
        document.body.appendChild(a);
        a.click();
        document.body.removeChild(a);
        URL.revokeObjectURL(url);
        
        errorDiv.textContent = ' Export completed successfully';
        errorDiv.style.background = '#e8f5e9';
        errorDiv.style.borderLeft = '4px solid #4CAF50';
        errorDiv.style.color = '#2e7d32';
        errorDiv.classList.add('show');
    } catch (error) {
        errorDiv.textContent = ' Export failed: ' + error.message;
        errorDiv.classList.add('show');
    }
}

window.onload = loadDashboard;
</script>
</body>
</html>"""

# ==================== FASTAPI APP ====================

app = FastAPI()

@app.get("/api/warranty-data")
async def get_warranty_data():
    try:
        if WARRANTY_DATA['credit_df'] is None:
            return {"credit": [], "debit": [], "arbitration": [], "currentMonth": [], "compensation": [], "prApproval": []}
        
        credit_records = WARRANTY_DATA['credit_df'].to_dict('records')
        debit_records = WARRANTY_DATA['debit_df'].to_dict('records')
        arbitration_records = WARRANTY_DATA['arbitration_df'].to_dict('records')
        current_month_records = (WARRANTY_DATA['current_month_df'].to_dict('records') if WARRANTY_DATA['current_month_df'] is not None else [])
        compensation_records = (WARRANTY_DATA['compensation_df'].to_dict('records') if WARRANTY_DATA['compensation_df'] is not None else [])
        pr_approval_records = (WARRANTY_DATA['pr_approval_df'].to_dict('records') if WARRANTY_DATA['pr_approval_df'] is not None else [])
        
        for records in [credit_records, debit_records, arbitration_records, current_month_records, compensation_records, pr_approval_records]:
            for record in records:
                for key in record:
                    if pd.isna(record[key]):
                        record[key] = 0
        
        return {
            "credit": credit_records, "debit": debit_records, "arbitration": arbitration_records,
            "currentMonth": current_month_records, "compensation": compensation_records, "prApproval": pr_approval_records
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail="Error loading data")

@app.post("/api/export-to-excel")
async def export_to_excel(request: Request):
    try:
        body = await request.json()
        selected_division = body.get('division', 'All')
        export_type = body.get('type', 'credit')
        
        print(f"\nEXPORT: {selected_division} - {export_type}")
        
        header_fill = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        wb = Workbook()
        wb.remove(wb.active)
        
        # ============ CREDIT ============
        if export_type == 'credit':
            df = WARRANTY_DATA['credit_df']
            if df is None or df.empty:
                raise HTTPException(status_code=500, detail="No data")
            
            if selected_division != 'All':
                df_export = df[df['Division'] == selected_division].copy()
                gt = df[df['Division'] == 'Grand Total']
                if not gt.empty:
                    df_export = pd.concat([df_export, gt], ignore_index=True)
            else:
                df_export = df.copy()
            
            ws = wb.create_sheet("Summary")
            style_worksheet(ws, df_export, header_fill, header_font, border)
        
        # ============ DEBIT ============
        elif export_type == 'debit':
            df = WARRANTY_DATA['debit_df']
            if df is None or df.empty:
                raise HTTPException(status_code=500, detail="No data")
            
            if selected_division != 'All':
                df_export = df[df['Division'] == selected_division].copy()
                gt = df[df['Division'] == 'Grand Total']
                if not gt.empty:
                    df_export = pd.concat([df_export, gt], ignore_index=True)
            else:
                df_export = df.copy()
            
            ws = wb.create_sheet("Summary")
            style_worksheet(ws, df_export, header_fill, header_font, border)
        
        # ============ ARBITRATION (3 SHEETS) ============
        elif export_type == 'arbitration':
            df = WARRANTY_DATA['arbitration_df']
            source_df = WARRANTY_DATA['source_df']
            
            if df is None or df.empty:
                raise HTTPException(status_code=500, detail="No data")
            
            # Sheet 1: Summary
            if selected_division != 'All':
                df_export = df[df['Division'] == selected_division].copy()
                gt = df[df['Division'] == 'Grand Total']
                if not gt.empty:
                    df_export = pd.concat([df_export, gt], ignore_index=True)
            else:
                df_export = df.copy()
            
            ws1 = wb.create_sheet("Summary")
            style_worksheet(ws1, df_export, header_fill, header_font, border)
            
            # Sheet 2 & 3: Claimed & Not Claimed
            if source_df is not None:
                def is_arbitration(value):
                    if pd.isna(value): return False
                    value = str(value).strip().upper()
                    return value.startswith('ARB') and value != 'NAN' and value != ''
                
                dealer_mapping = {
                    'AMT': 'AMRAVATI', 'CHA': 'CHAUFULA_SZZ', 'CHI': 'CHIKHALI',
                    'KOL': 'KOLHAPUR_WS', 'HO': 'NAGPUR_KAMPTHEE ROAD',
                    'CITY': 'NAGPUR_WARDHAMAN NGR', 'SHI': 'SHIKRAPUR_SZS',
                    'WAG': 'WAGHOLI', 'YAT': 'YAVATMAL', 'CQ': 'NAGPUR_WARDHAMAN NGR_CQ'
                }
                
                if selected_division != 'All':
                    dealer_location = dealer_mapping.get(selected_division)
                    filtered_source = source_df[source_df['Dealer Location'] == dealer_location].copy() if dealer_location else source_df.copy()
                else:
                    filtered_source = source_df.copy()
                
                # Claimed
                claimed_df = filtered_source[filtered_source['Claim arbitration ID'].apply(is_arbitration)][
                    ['Fiscal Month', 'Dealer Location', 'Debit Note Amount', 'Claim arbitration ID']
                ].copy()
                
                if not claimed_df.empty:
                    ws2 = wb.create_sheet("Arbitration Claimed")
                    style_worksheet(ws2, claimed_df, header_fill, header_font, border)
                
                # Not Claimed
                not_claimed_df = filtered_source[~filtered_source['Claim arbitration ID'].apply(is_arbitration)][
                    ['Fiscal Month', 'Dealer Location', 'Debit Note Amount']
                ].copy()
                
                if not not_claimed_df.empty:
                    ws3 = wb.create_sheet("Arbitration Not Claimed")
                    style_worksheet(ws3, not_claimed_df, header_fill, header_font, border)
        
        # ============ CURRENT MONTH (3 SHEETS) ============
        elif export_type == 'currentmonth':
            df = WARRANTY_DATA['current_month_df']
            source_df = WARRANTY_DATA['current_month_source_df']
            
            if df is None or df.empty:
                raise HTTPException(status_code=500, detail="No data")
            
            # Sheet 1: Summary
            if selected_division != 'All':
                df_export = df[df['Division'] == selected_division].copy()
                gt = df[df['Division'] == 'Grand Total']
                if not gt.empty:
                    df_export = pd.concat([df_export, gt], ignore_index=True)
            else:
                df_export = df.copy()
            
            ws1 = wb.create_sheet("Summary")
            style_worksheet(ws1, df_export, header_fill, header_font, border)
            
            # Sheet 2 & 3: Details
            if source_df is not None:
                if selected_division != 'All':
                    spares_df = source_df[(source_df['Division'] == selected_division) & 
                                         (source_df['Pending Claims Spares'].notna())].copy()
                    labour_df = source_df[(source_df['Division'] == selected_division) & 
                                         (source_df['Pending Claims Labour'].notna())].copy()
                else:
                    spares_df = source_df[source_df['Pending Claims Spares'].notna()].copy()
                    labour_df = source_df[source_df['Pending Claims Labour'].notna()].copy()
                
                if not spares_df.empty:
                    ws2 = wb.create_sheet("Spares Claimed Details")
                    style_worksheet(ws2, spares_df, header_fill, header_font, border)
                
                if not labour_df.empty:
                    ws3 = wb.create_sheet("Labour Claimed Details")
                    style_worksheet(ws3, labour_df, header_fill, header_font, border)
        
        # ============ COMPENSATION (TAT FROM RO BILL DATE) ============
        elif export_type == 'compensation':
            summary_df = WARRANTY_DATA['compensation_df']
            source_df = WARRANTY_DATA['compensation_source_df']
            
            if summary_df is None or summary_df.empty:
                raise HTTPException(status_code=500, detail="No data")
            
            # Sheet 1: Summary
            if selected_division != 'All':
                summary_export = summary_df[summary_df['Division'] == selected_division].copy()
                gt = summary_df[summary_df['Division'] == 'Grand Total']
                if not gt.empty:
                    summary_export = pd.concat([summary_export, gt], ignore_index=True)
            else:
                summary_export = summary_df.copy()
            
            ws_summary = wb.create_sheet("Summary")
            style_worksheet(ws_summary, summary_export, header_fill, header_font, border)
            
            # Sheet 2: Details with TAT (RO BILL DATE TO CURRENT DATE)
            if source_df is not None and not source_df.empty:
                if selected_division != 'All':
                    detail_df = source_df[source_df['Division'] == selected_division].copy()
                else:
                    detail_df = source_df.copy()
                
                if not detail_df.empty:
                    required_cols = [
                        'Division', 'RO Id.', 'Registration No.', 'Chassis No.', 'Model Group',
                        'RO Date', 'RO Bill Date', 'Claim Amount', 'Claim Date',
                        'Request No.', 'Request Date', 'Request Status', 
                        'Claim Approved Amt.', 'No. of Days'
                    ]
                    
                    available_cols = [col for col in required_cols if col in detail_df.columns]
                    detail_df = detail_df[available_cols].copy()
                    
                    # REPLACE "No. of Days" with TAT from RO BILL DATE
                    if 'RO Bill Date' in detail_df.columns:
                        def calculate_tat_from_bill(bill_date):
                            try:
                                if pd.isna(bill_date):
                                    return 0
                                bill_dt = pd.to_datetime(bill_date, errors='coerce')
                                if pd.isna(bill_dt):
                                    return 0
                                tat = (datetime.now().date() - bill_dt.date()).days
                                return max(0, tat)
                            except:
                                return 0
                        
                        # Replace "No. of Days" column with TAT
                        if 'No. of Days' in detail_df.columns:
                            detail_df['No. of Days'] = detail_df['RO Bill Date'].apply(calculate_tat_from_bill)
                    
                    # Convert dates
                    date_columns = ['RO Date', 'RO Bill Date', 'Claim Date', 'Request Date']
                    for col in date_columns:
                        if col in detail_df.columns:
                            detail_df[col] = pd.to_datetime(detail_df[col], errors='coerce')
                    
                    # Sort
                    if 'Request Date' in detail_df.columns:
                        detail_df = detail_df.sort_values('Request Date', ascending=False, na_position='last')
                    
                    ws_details = wb.create_sheet("Details")
                    style_worksheet(ws_details, detail_df, header_fill, header_font, border)
        
        # ============ PR APPROVAL (SUMMARY + DETAILS) ============
        elif export_type == 'pr_approval':
            df = WARRANTY_DATA['pr_approval_df']
            source_df = WARRANTY_DATA['pr_approval_source_df']
            
            if df is None or df.empty:
                raise HTTPException(status_code=500, detail="No data")
            
            # Sheet 1: Summary
            if selected_division != 'All':
                df_export = df[df['Division'] == selected_division].copy()
                gt = df[df['Division'] == 'Grand Total']
                if not gt.empty:
                    df_export = pd.concat([df_export, gt], ignore_index=True)
            else:
                df_export = df.copy()
            
            ws1 = wb.create_sheet("Summary")
            style_worksheet(ws1, df_export, header_fill, header_font, border)
            
            # Sheet 2: Details (ALL COLUMNS FROM YOUR OLD CODE)
            if source_df is not None and not source_df.empty:
                if selected_division != 'All':
                    detail_df = source_df[source_df['Division'] == selected_division].copy()
                else:
                    detail_df = source_df.copy()
                
                if not detail_df.empty:
                    ws2 = wb.create_sheet("Details")
                    style_worksheet(ws2, detail_df, header_fill, header_font, border)
        
        # Save
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"{selected_division}_{export_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        print(f" Export ready: {filename}\n")
        
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        print(f" Error: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/")
async def root():
    return HTMLResponse(content=DASHBOARD_HTML)

# ==================== STARTUP ====================

print("Processing warranty data...\n")
print("1. Processing Warranty Debit...")
WARRANTY_DATA['credit_df'], WARRANTY_DATA['debit_df'], WARRANTY_DATA['arbitration_df'], WARRANTY_DATA['source_df'] = process_warranty_data()

print("2. Processing Current Month...")
WARRANTY_DATA['current_month_df'], WARRANTY_DATA['current_month_source_df'] = process_current_month_warranty()

print("3. Processing Compensation...")
WARRANTY_DATA['compensation_df'], WARRANTY_DATA['compensation_source_df'] = process_compensation_claim()

print("4. Processing PR Approval...")
WARRANTY_DATA['pr_approval_df'], WARRANTY_DATA['pr_approval_source_df'] = process_pr_approval()

if __name__ == "__main__":
    port = int(os.getenv('PORT', 8001))
    
    print("\n" + "="*100)
    print(" READY - WARRANTY MANAGEMENT SYSTEM")
    print("="*100)
    print(f"Access: http://localhost:{port}")
    print("\nEXPORT OPTIONS (CORRECTED LABELS):")
    print("   Credit (no '2 sheets')")
    print("   Debit (no '2 sheets')")
    print("   Arbitration (3 sheets: Summary + Claimed + Not Claimed)")
    print("   Current Month (3 sheets: Summary + Spares Details + Labour Details)")
    print("   Compensation (TAT from RO Bill Date)")
    print("   PR Approval (Summary + Details)")
    print("="*100 + "\n")
    
    uvicorn.run(app, host="0.0.0.0", port=port)
