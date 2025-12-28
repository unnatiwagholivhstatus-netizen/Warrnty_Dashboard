import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import uvicorn
from fastapi import FastAPI, Request, HTTPException, Cookie
from fastapi.responses import HTMLResponse, FileResponse, RedirectResponse, JSONResponse, StreamingResponse
import os
import socket
from typing import Optional
import sys
from functools import lru_cache
import hashlib
import secrets
import string
from PIL import Image, ImageDraw, ImageFont
import io
import base64
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

# ==================== WARRANTY DATA PROCESSING ====================

WARRANTY_DATA = {
    'credit_df': None,
    'debit_df': None,
    'arbitration_df': None,
    'source_df': None,
    'current_month_df': None,
    'current_month_source_df': None,
    'compensation_df': None,
    'compensation_source_df': None,
    'pr_approval_df': None,
    'pr_approval_source_df': None
}

def find_data_file(filename):
    """Find data file in multiple possible locations"""
    possible_paths = [
        f"/mnt/data/{filename}",
        filename,
        f"./{filename}",
        f"Data/{filename}",
        f"data/{filename}",
    ]
    
    for path in possible_paths:
        if os.path.exists(path):
            print(f"  Found: {filename} at {path}")
            return path
    
    print(f"  WARNING: {filename} not found. Checked: {possible_paths}")
    return None

def process_pr_approval():
    """Process PR Approval data and return summary dataframe"""
    #  FIXED: Correct file path pointing to Pr_Approval_Claims_Merged.xlsx
    input_path = find_data_file('Pr_Approval_Claims_Merged.xlsx')
    
    if input_path is None:
        print("  PR Approval file not found - returning empty data")
        return None, None
    
    try:
        # Load the data - read first sheet
        df = pd.read_excel(input_path)
        print("  PR Approval data loaded successfully")
        print(f"  Available columns: {df.columns.tolist()[:10]}...")
        print(f"  Total rows in source data: {len(df)}")

        # Required columns for the summary table
        summary_columns = [
            'Division', 'PA Request No.', 'PA Date', 'Request Type', 'App. Claim Amt from M&M'
        ]
        
        # Check which columns exist
        available_summary_columns = [col for col in summary_columns if col in df.columns]
        missing_columns = [col for col in summary_columns if col not in df.columns]
        
        if missing_columns:
            print(f" Missing columns in PR Approval: {missing_columns}")
            print(f" Available columns: {df.columns.tolist()}")
        
        if not available_summary_columns:
            print(f" No required columns found in PR Approval file")
            return None, None

        # Select only available summary columns for display
        df_summary_display = df[available_summary_columns].copy()
        
        # Clean the Division column
        if 'Division' in df_summary_display.columns:
            df_summary_display['Division'] = df_summary_display['Division'].astype(str).str.strip()
            df_summary_display = df_summary_display[df_summary_display['Division'].notna() & 
                                                      (df_summary_display['Division'] != '') & 
                                                      (df_summary_display['Division'] != 'nan')]
        
        # Clean numeric columns
        if 'App. Claim Amt from M&M' in df_summary_display.columns:
            df_summary_display['App. Claim Amt from M&M'] = pd.to_numeric(
                df_summary_display['App. Claim Amt from M&M'], errors='coerce'
            ).fillna(0)
        
        # Prepare summary by division
        summary_data = []
        
        if 'Division' in df_summary_display.columns:
            for division in sorted(df_summary_display['Division'].unique()):
                div_data = df_summary_display[df_summary_display['Division'] == division]
                
                summary_row = {'Division': division}
                
                # Count of requests
                summary_row['Total Requests'] = len(div_data)
                
                # Sum of App. Claim Amt from M&M
                if 'App. Claim Amt from M&M' in df_summary_display.columns:
                    summary_row['Total Approved Amount'] = div_data['App. Claim Amt from M&M'].sum()
                
                # Count by Request Type if available
                if 'Request Type' in df_summary_display.columns:
                    request_types = div_data['Request Type'].value_counts().to_dict()
                    for req_type, count in request_types.items():
                        if pd.notna(req_type) and str(req_type).strip() != '':
                            summary_row[f'{req_type} Count'] = count
                
                summary_data.append(summary_row)
            
            # Create summary dataframe
            summary_df = pd.DataFrame(summary_data)
            
            # Add Grand Total row
            grand_total = {'Division': 'Grand Total'}
            
            for col in summary_df.columns:
                if col != 'Division':
                    if summary_df[col].dtype in ['int64', 'float64']:
                        grand_total[col] = summary_df[col].sum()
            
            summary_df = pd.concat([summary_df, pd.DataFrame([grand_total])], ignore_index=True)
        else:
            summary_df = pd.DataFrame()

        print("\n PR Approval processing completed successfully")
        if not summary_df.empty:
            print(f"  Total Requests: {len(df_summary_display)}")
            if 'App. Claim Amt from M&M' in df_summary_display.columns:
                print(f"  Total Approved Amount: {df_summary_display['App. Claim Amt from M&M'].sum():,.2f}")
        
        # Return summary and complete source dataframe for export
        return summary_df, df

    except FileNotFoundError:
        print(f" PR Approval file not found: {input_path}")
        return None, None
    except Exception as e:
        import traceback
        print(f" Error processing PR Approval data: {e}")
        traceback.print_exc()
        return None, None

def process_compensation_claim():
    """Process compensation claim data and return summary dataframe"""
    input_path = find_data_file('Transit_Claims_Merged.xlsx')
    
    if input_path is None:
        print("  Compensation Claim file not found - returning empty data")
        return None, None
    
    try:
        # Load the data - read first sheet
        df = pd.read_excel(input_path)
        print(" Compensation Claim data loaded successfully")
        print(f"  Available columns: {df.columns.tolist()[:10]}...")
        print(f"  Total rows in source data: {len(df)}")

        # Required columns for the table
        required_columns = [
            'Division', 'RO Id.', 'Registration No.', 'RO Date', 'RO Bill Date',
            'Chassis No.', 'Model Group', 'Claim Amount', 'Request Status',
            'Claim Approved Amt.', 'No. of Days'
        ]
        
        # Check which columns exist
        available_columns = [col for col in required_columns if col in df.columns]
        missing_columns = [col for col in required_columns if col not in df.columns]
        
        if missing_columns:
            print(f" Missing columns in Compensation Claim: {missing_columns}")
            print(f" Available columns: {df.columns.tolist()}")
        
        if not available_columns:
            print(f" No required columns found in Compensation Claim file")
            return None, None

        # Select only available columns
        df_filtered = df[available_columns].copy()
        
        # Clean the Division column
        if 'Division' in df_filtered.columns:
            df_filtered['Division'] = df_filtered['Division'].astype(str).str.strip()
            df_filtered = df_filtered[df_filtered['Division'].notna() & (df_filtered['Division'] != '') & (df_filtered['Division'] != 'nan')]
        
        # Format RO Id with "RO" prefix if column exists
        if 'RO Id.' in df_filtered.columns:
            def format_ro_id(x):
                if pd.isna(x) or str(x).strip() == '':
                    return ''
                try:
                    return f"RO{str(int(float(x)))}"
                except (ValueError, TypeError):
                    value_str = str(x).strip()
                    if not value_str.startswith('RO'):
                        return f"RO{value_str}"
                    return value_str
            
            df_filtered['RO Id.'] = df_filtered['RO Id.'].apply(format_ro_id)
        
        # Clean numeric columns
        numeric_cols = ['Claim Amount', 'Claim Approved Amt.', 'No. of Days']
        for col in numeric_cols:
            if col in df_filtered.columns:
                df_filtered[col] = pd.to_numeric(df_filtered[col], errors='coerce').fillna(0)
        
        # Prepare summary by division
        summary_data = []
        
        if 'Division' in df_filtered.columns:
            for division in sorted(df_filtered['Division'].unique()):
                div_data = df_filtered[df_filtered['Division'] == division]
                
                summary_row = {'Division': division}
                
                # Count of claims
                summary_row['Total Claims'] = len(div_data)
                
                # Sum of Claim Amount
                if 'Claim Amount' in df_filtered.columns:
                    summary_row['Total Claim Amount'] = div_data['Claim Amount'].sum()
                
                # Sum of Claim Approved Amount
                if 'Claim Approved Amt.' in df_filtered.columns:
                    summary_row['Total Approved Amount'] = div_data['Claim Approved Amt.'].sum()
                
                # Average No. of Days
                if 'No. of Days' in df_filtered.columns:
                    summary_row['Avg No. of Days'] = div_data['No. of Days'].mean()
                
                summary_data.append(summary_row)
            
            # Create summary dataframe
            summary_df = pd.DataFrame(summary_data)
            
            # Add Grand Total row
            grand_total = {'Division': 'Grand Total'}
            
            if 'Total Claims' in summary_df.columns:
                grand_total['Total Claims'] = summary_df['Total Claims'].sum()
            
            if 'Total Claim Amount' in summary_df.columns:
                grand_total['Total Claim Amount'] = summary_df['Total Claim Amount'].sum()
            
            if 'Total Approved Amount' in summary_df.columns:
                grand_total['Total Approved Amount'] = summary_df['Total Approved Amount'].sum()
            
            if 'Avg No. of Days' in summary_df.columns:
                grand_total['Avg No. of Days'] = summary_df['Avg No. of Days'].mean()
            
            summary_df = pd.concat([summary_df, pd.DataFrame([grand_total])], ignore_index=True)
        else:
            summary_df = pd.DataFrame()

        print("\n Compensation Claim processing completed successfully")
        if not summary_df.empty:
            print(f"  Total Claims: {len(df_filtered)}")
            if 'Claim Amount' in df_filtered.columns:
                print(f"  Total Claim Amount: {df_filtered['Claim Amount'].sum():,.2f}")
        
        return summary_df, df_filtered

    except FileNotFoundError:
        print(f" Compensation Claim file not found: {input_path}")
        return None, None
    except Exception as e:
        import traceback
        print(f" Error processing compensation claim data: {e}")
        traceback.print_exc()
        return None, None

def process_current_month_warranty():
    """Process current month warranty data and return summary dataframe"""
    input_path = find_data_file('Pending Warranty Claim Details.xlsx')
    
    if input_path is None:
        print("  Current Month Warranty file not found - returning empty data")
        return None, None
    
    try:
        # Load the data - sheet name is "Pending Warranty Claim Details"
        df = pd.read_excel(input_path, sheet_name='Pending Warranty Claim Details')
        print(" Current Month Warranty data loaded successfully")
        print(f"  Available columns: {df.columns.tolist()[:10]}...")
        print(f"  Total rows in source data: {len(df)}")

        # Check if required columns exist
        required_columns = ['Division', 'Pending Claims Spares', 'Pending Claims Labour']
        missing_columns = [col for col in required_columns if col not in df.columns]
        
        if missing_columns:
            print(f" Missing columns in Current Month Warranty: {missing_columns}")
            print(f" Available columns: {df.columns.tolist()}")
            return None, None

        # Clean the Division column
        df['Division'] = df['Division'].astype(str).str.strip()
        
        # Remove any empty or NaN divisions
        df = df[df['Division'].notna() & (df['Division'] != '') & (df['Division'] != 'nan')]

        # Prepare summary by division
        summary_data = []
        
        for division in sorted(df['Division'].unique()):
            div_data = df[df['Division'] == division]
            
            # Count non-empty Pending Claims Spares
            spares_count = div_data['Pending Claims Spares'].notna().sum()
            
            # Count non-empty Pending Claims Labour
            labour_count = div_data['Pending Claims Labour'].notna().sum()
            
            summary_data.append({
                'Division': division,
                'Pending Claims Spares Count': spares_count,
                'Pending Claims Labour Count': labour_count,
                'Total Pending Claims': spares_count + labour_count
            })
        
        # Create summary dataframe
        summary_df = pd.DataFrame(summary_data)
        
        # Add Grand Total row
        grand_total = {
            'Division': 'Grand Total',
            'Pending Claims Spares Count': summary_df['Pending Claims Spares Count'].sum(),
            'Pending Claims Labour Count': summary_df['Pending Claims Labour Count'].sum(),
            'Total Pending Claims': summary_df['Total Pending Claims'].sum()
        }
        summary_df = pd.concat([summary_df, pd.DataFrame([grand_total])], ignore_index=True)

        print("\n Current Month Warranty processing completed successfully")
        print(f"  Total Pending Claims Spares: {grand_total['Pending Claims Spares Count']}")
        print(f"  Total Pending Claims Labour: {grand_total['Pending Claims Labour Count']}")
        
        return summary_df, df

    except FileNotFoundError:
        print(f" Current Month Warranty file not found: {input_path}")
        return None, None
    except Exception as e:
        import traceback
        print(f" Error processing current month warranty data: {e}")
        traceback.print_exc()
        return None, None

def process_warranty_data():
    """Process warranty data and return credit, debit, and arbitration dataframes"""
    input_path = find_data_file('Warranty Debit.xlsx')
    
    if input_path is None:
        print("  Warranty Debit file not found - returning empty data")
        return None, None, None, None
    
    try:
        # Load the data
        df = pd.read_excel(input_path, sheet_name='Sheet1')
        print(" Warranty data loaded successfully")
        print(f"  Available columns: {df.columns.tolist()[:5]}...")
        print(f"  Total rows in source data: {len(df)}")

        # Dealer location mapping
        dealer_mapping = {
            'AMRAVATI': 'AMT',
            'CHAUFULA_SZZ': 'CHA',
            'CHIKHALI': 'CHI',
            'KOLHAPUR_WS': 'KOL',
            'NAGPUR_KAMPTHEE ROAD': 'HO',
            'NAGPUR_WARDHAMAN NGR': 'CITY',
            'SHIKRAPUR_SZS': 'SHI',
            'WAGHOLI': 'WAG',
            'YAVATMAL': 'YAT',
            'NAGPUR_WARDHAMAN NGR_CQ': 'CQ'
        }

        # Clean numeric columns
        numeric_columns = ['Total Claim Amount', 'Credit Note Amount', 'Debit Note Amount']
        for col in numeric_columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

        print(f"\n  Summary:")
        print(f"    Total Credit Note: {df['Credit Note Amount'].sum():,.2f}")
        print(f"    Total Debit Note: {df['Debit Note Amount'].sum():,.2f}")

        # Apply dealer mapping
        df['Dealer_Code'] = df['Dealer Location'].map(dealer_mapping).fillna(df['Dealer Location'])

        # Extract month from 'Fiscal Month'
        df['Month'] = df['Fiscal Month'].astype(str).str.strip().str[:3]

        # Ensure 'Claim arbitration ID' is clean
        df['Claim arbitration ID'] = df['Claim arbitration ID'].astype(str).replace('nan', '').replace('', np.nan)

        # Prepare result table
        dealers = sorted(df['Dealer_Code'].unique())
        months = ['Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

        # 1. CREDIT NOTE TABLE
        credit_df = pd.DataFrame({'Division': dealers})
        print("\n  Processing Credit Note Amounts...")
        for month in months:
            month_data = df[df['Month'] == month]
            if not month_data.empty:
                summary = month_data.groupby('Dealer_Code')['Credit Note Amount'].sum().reset_index()
                summary.columns = ['Division', f'Credit Note {month}']
                credit_df = credit_df.merge(summary, on='Division', how='left')
                print(f"    {month}: {month_data['Credit Note Amount'].sum():,.2f}")
            else:
                credit_df[f'Credit Note {month}'] = 0
        
        credit_df = credit_df.fillna(0)
        credit_columns = [f'Credit Note {month}' for month in months]
        credit_df['Total Credit'] = credit_df[credit_columns].sum(axis=1)
        
        # Add Grand Total row
        grand_total_credit = {'Division': 'Grand Total'}
        for col in credit_df.columns[1:]:
            grand_total_credit[col] = credit_df[col].sum()
        credit_df = pd.concat([credit_df, pd.DataFrame([grand_total_credit])], ignore_index=True)

        # 2. DEBIT NOTE TABLE
        debit_df = pd.DataFrame({'Division': dealers})
        print("\n  Processing Debit Note Amounts...")
        for month in months:
            month_data = df[df['Month'] == month]
            if not month_data.empty:
                summary = month_data.groupby('Dealer_Code')['Debit Note Amount'].sum().reset_index()
                summary.columns = ['Division', f'Debit Note {month}']
                debit_df = debit_df.merge(summary, on='Division', how='left')
                print(f"    {month}: {month_data['Debit Note Amount'].sum():,.2f}")
            else:
                debit_df[f'Debit Note {month}'] = 0
        
        debit_df = debit_df.fillna(0)
        debit_columns = [f'Debit Note {month}' for month in months]
        debit_df['Total Debit'] = debit_df[debit_columns].sum(axis=1)
        
        # Add Grand Total row
        grand_total_debit = {'Division': 'Grand Total'}
        for col in debit_df.columns[1:]:
            grand_total_debit[col] = debit_df[col].sum()
        debit_df = pd.concat([debit_df, pd.DataFrame([grand_total_debit])], ignore_index=True)

        # 3. CLAIM ARBITRATION TABLE
        arbitration_df = pd.DataFrame({'Division': dealers})
        print("\n  Processing Claim Arbitration...")
        
        def is_arbitration(value):
            if pd.isna(value): return False
            value = str(value).strip().upper()
            return value.startswith('ARB') and value != 'NAN'

        for month in months:
            month_data = df[df['Month'] == month].copy()
            month_data['Is_ARB'] = month_data['Claim arbitration ID'].apply(is_arbitration)
            month_data['Arbitration_Amount'] = month_data.apply(
                lambda row: row['Debit Note Amount'] if row['Is_ARB'] else 0,
                axis=1
            )
            arb_summary = month_data.groupby('Dealer_Code')['Arbitration_Amount'].sum().reset_index()
            arb_summary.columns = ['Division', f'Claim Arbitration {month}']
            arbitration_df = arbitration_df.merge(arb_summary, on='Division', how='left')
            print(f"    {month}: {month_data['Arbitration_Amount'].sum():,.2f}")
        
        arbitration_df = arbitration_df.fillna(0)
        
        # Calculate Pending Claim Arbitration
        arbitration_cols = [f'Claim Arbitration {m}' for m in months]
        
        # Get Total Debit for each dealer (without Grand Total)
        total_debit_by_dealer = debit_df[debit_df['Division'] != 'Grand Total'][['Division', 'Total Debit']].copy()
        arbitration_df = arbitration_df.merge(total_debit_by_dealer, on='Division', how='left')
        
        arbitration_df['Pending Claim Arbitration'] = (
            arbitration_df['Total Debit'] - arbitration_df[arbitration_cols].sum(axis=1)
        )
        
        # Remove Total Debit column
        arbitration_df = arbitration_df.drop('Total Debit', axis=1)
        
        # Add Grand Total row
        grand_total_arb = {'Division': 'Grand Total'}
        for col in arbitration_df.columns[1:]:
            grand_total_arb[col] = arbitration_df[col].sum()
        arbitration_df = pd.concat([arbitration_df, pd.DataFrame([grand_total_arb])], ignore_index=True)

        print("\n Warranty data processing completed successfully")
        return credit_df, debit_df, arbitration_df, df

    except FileNotFoundError:
        print(f" Warranty file not found: {input_path}")
        return None, None, None, None
    except Exception as e:
        import traceback
        print(f" Error processing warranty data: {e}")
        traceback.print_exc()
        return None, None, None, None

# ==================== IMAGE HANDLING ====================

def get_mahindra_images():
    """Load Mahindra vehicle images from the folder"""
    image_folder = r"D:\Power BI New\Warranty Debit\Image"
    images = []
    branding_images = []
    vehicle_images = []
    
    if os.path.exists(image_folder):
        try:
            for file in os.listdir(image_folder):
                if file.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp')):
                    image_path = os.path.join(image_folder, file)
                    try:
                        with open(image_path, 'rb') as img_file:
                            img_data = base64.b64encode(img_file.read()).decode()
                            img_dict = {
                                'name': file,
                                'data': img_data,
                                'path': image_path
                            }
                            
                            file_lower = file.lower()
                            if 'mahindra' in file_lower or 'logo' in file_lower or 'hero' in file_lower:
                                branding_images.append(img_dict)
                                print(f"   Loaded Branding: {file}")
                            else:
                                vehicle_images.append(img_dict)
                                print(f"   Loaded Vehicle: {file}")
                    except Exception as e:
                        print(f"   Could not load {file}: {e}")
        except Exception as e:
            print(f" Error reading image folder: {e}")
    else:
        print(f" Image folder not found: {image_folder}")
    
    images = branding_images + vehicle_images
    return images

print("Loading Mahindra vehicle images...")
MAHINDRA_IMAGES = get_mahindra_images()
print(f" Loaded {len(MAHINDRA_IMAGES)} vehicle images\n")

# ==================== AUTHENTICATION SETUP ====================

def update_user_password_in_excel(user_id: str, new_password: str):
    """Update user password in UserID.xlsx Excel file"""
    try:
        user_file = r"D:\Power BI New\Warranty Debit\UserID.xlsx"
        
        if not os.path.exists(user_file):
            print(f" ERROR: User file not found: {user_file}")
            return False
        
        # Read the Excel file
        df = pd.read_excel(user_file)
        
        # Find the row with matching User ID
        mask = df['User ID'].apply(lambda x: str(int(float(x))) == str(user_id) if pd.notna(x) else False)
        
        if not mask.any():
            print(f" User ID {user_id} not found in file")
            return False
        
        # Update the password
        df.loc[mask, 'Password'] = new_password
        
        # Write back to Excel
        df.to_excel(user_file, index=False)
        
        print(f" Password updated in Excel for User ID: {user_id}")
        return True
        
    except Exception as e:
        print(f" Error updating password in Excel: {e}")
        import traceback
        traceback.print_exc()
        return False

def load_user_credentials():
    """Load user credentials from UserID.xlsx"""
    try:
        user_file = r"D:\Power BI New\Warranty Debit\UserID.xlsx"
        if not os.path.exists(user_file):
            print(f" ERROR: User file not found: {user_file}")
            return {}
        
        df = pd.read_excel(user_file)
        print(f" Loaded user file from {user_file}")
        print(f"  Total rows: {len(df)}")
        
        credentials = {}
        for idx, row in df.iterrows():
            try:
                # Handle User ID - convert float to int to string
                user_id = None
                if 'User ID' in df.columns:
                    uid_value = row['User ID']
                    # Check if it's NaN
                    if pd.isna(uid_value):
                        continue
                    # Convert float to int, then to string
                    user_id = str(int(float(uid_value)))
                
                # Handle Password
                password = None
                if 'Password' in df.columns:
                    pwd_value = row['Password']
                    # Check if it's NaN
                    if pd.isna(pwd_value):
                        continue
                    password = str(pwd_value).strip()
                
                if user_id and password:
                    credentials[user_id] = password
            except Exception as e:
                continue
        
        print(f" Successfully loaded {len(credentials)} valid user credentials")
        print(f"  Available User IDs: {list(credentials.keys())}\n")
        return credentials
    except Exception as e:
        import traceback
        print(f" ERROR loading credentials: {e}")
        traceback.print_exc()
        return {}

USER_CREDENTIALS = load_user_credentials()
SESSIONS = {}

class CaptchaGenerator:
    """Generate simple CAPTCHA images"""
    
    @staticmethod
    def generate_captcha(length=6):
        """Generate random captcha string and image"""
        allowed_chars = 'ABCDEFGHJKLMNPQRSTUVWXYZ123456789'
        captcha_text = ''.join(secrets.choice(allowed_chars) for _ in range(length))
        
        width, height = 500, 150
        image = Image.new('RGB', (width, height), color='white')
        draw = ImageDraw.Draw(image)
        
        for _ in range(5):
            x1 = secrets.randbelow(width)
            y1 = secrets.randbelow(height)
            x2 = secrets.randbelow(width)
            y2 = secrets.randbelow(height)
            draw.line((x1, y1, x2, y2), fill='lightgray', width=1)
        
        try:
            font_size = 80
            font = ImageFont.truetype("C:\\Windows\\Fonts\\arial.ttf", font_size)
        except:
            try:
                font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 80)
            except:
                font = ImageFont.load_default()
        
        x_offset = 15
        for i, char in enumerate(captcha_text):
            y_offset = np.random.randint(15, 50)
            draw.text((x_offset + i * 70, y_offset), char, fill='#FF8C00', font=font)
        
        for _ in range(50):
            x = secrets.randbelow(width)
            y = secrets.randbelow(height)
            draw.point((x, y), fill='#FFD699')
        
        img_io = io.BytesIO()
        image.save(img_io, 'PNG')
        img_io.seek(0)
        img_base64 = base64.b64encode(img_io.getvalue()).decode()
        
        return captcha_text, f"data:image/png;base64,{img_base64}"

def create_session(user_id):
    """Create a new session for user"""
    session_id = secrets.token_hex(16)
    SESSIONS[session_id] = {
        'user_id': user_id,
        'created_at': datetime.now(),
        'last_activity': datetime.now()
    }
    return session_id

def verify_session(session_id):
    """Verify if session is valid"""
    if session_id not in SESSIONS:
        return None
    
    session = SESSIONS[session_id]
    if (datetime.now() - session['last_activity']).total_seconds() > 8 * 3600:
        del SESSIONS[session_id]
        return None
    
    session['last_activity'] = datetime.now()
    return session['user_id']

# ==================== LOGIN PAGE HTML ====================

LOGIN_PAGE = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Warranty Management System - Login</title>
    <style>
        * {
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }
        
        body {
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: linear-gradient(135deg, #FF8C00 0%, #FF6B35 100%);
            height: 100vh;
            width: 100vw;
            display: flex;
            justify-content: center;
            align-items: center;
            padding: 0;
            margin: 0;
            overflow: hidden;
        }
        
        .login-wrapper {
            background: white;
            border-radius: 20px;
            box-shadow: 0 20px 60px rgba(0, 0, 0, 0.3);
            overflow: hidden;
            width: 95vw;
            height: 95vh;
            max-width: 1400px;
            max-height: 800px;
            display: grid;
            grid-template-columns: 1fr 1fr;
        }
        
        .login-left {
            padding: 30px;
            display: flex;
            flex-direction: column;
            justify-content: center;
            background: white;
            overflow: hidden;
            height: 100%;
        }
        
        .login-right {
            background: linear-gradient(135deg, #FF8C00 0%, #FF6B35 100%);
            padding: 25px;
            display: flex;
            flex-direction: column;
            justify-content: center;
            align-items: center;
            color: white;
            overflow: hidden;
            height: 100%;
        }
        
        .logo-section {
            text-align: center;
            margin-bottom: 15px;
        }
        
        .logo-section h1 {
            font-size: 28px;
            color: #333;
            margin-bottom: 8px;
        }
        
        .login-form {
            display: flex;
            flex-direction: column;
            gap: 20px;
        }
        
        .form-group {
            display: flex;
            flex-direction: column;
            gap: 8px;
        }
        
        .form-group label {
            font-weight: 600;
            color: #333;
            font-size: 14px;
        }
        
        .form-group input {
            padding: 12px;
            border: 2px solid #e0e0e0;
            border-radius: 8px;
            font-size: 14px;
            transition: all 0.3s ease;
        }
        
        .form-group input:focus {
            outline: none;
            border-color: #FF8C00;
            box-shadow: 0 0 8px rgba(255, 140, 0, 0.2);
        }
        
        .login-btn {
            background: linear-gradient(135deg, #FF8C00 0%, #FF6B35 100%);
            color: white;
            border: none;
            padding: 12px;
            border-radius: 8px;
            font-weight: 600;
            cursor: pointer;
            transition: all 0.3s ease;
            margin-top: 10px;
        }
        
        .login-btn:hover {
            transform: translateY(-2px);
            box-shadow: 0 5px 15px rgba(255, 140, 0, 0.3);
        }
        
        .error-message {
            color: #c62828;
            font-size: 13px;
            margin-top: -15px;
            display: none;
        }
        
        .error-message.show {
            display: block;
        }
        
        .captcha-section {
            margin-top: 15px;
            padding: 10px;
            background: #f5f5f5;
            border-radius: 8px;
        }
        
        .captcha-image {
            width: 100%;
            height: auto;
            margin-bottom: 10px;
            border-radius: 4px;
        }
        
        .right-content {
            text-align: center;
        }
        
        .right-content h2 {
            font-size: 32px;
            margin-bottom: 20px;
            text-shadow: 2px 2px 4px rgba(0, 0, 0, 0.2);
        }
        
        .right-content p {
            font-size: 16px;
            line-height: 1.6;
            opacity: 0.95;
        }
    </style>
</head>
<body>
    <div class="login-wrapper">
        <div class="login-left">
            <div class="logo-section">
                <h1>Unnati Motors Warranty Management</h1>
                <p style="color: #666; font-size: 14px;">Mahindra All Division Warranty Overview Dashboard</p>
                <p style="color: #666; font-size: 14px;">Enter your credentials to access the warranty dashboard</p>
            </div>
           
            <form class="login-form" onsubmit="handleLogin(event)">
                <div class="form-group">
                    <label for="userId">User ID</label>
                    <input type="text" id="userId" name="userId" placeholder="Enter your User ID" required>
                </div>
                
                <div class="form-group">
                    <label for="password">Password</label>
                    <input type="password" id="password" name="password" placeholder="Enter your password" required>
                </div>
                
                <div class="captcha-section">
                    <img id="captchaImage" class="captcha-image" src="" alt="CAPTCHA">
                    <input type="text" id="captchaInput" placeholder="Enter CAPTCHA" required style="width: 100%; padding: 8px; border: 2px solid #e0e0e0; border-radius: 4px;">
                </div>
                
                <div class="error-message" id="errorMessage"></div>
                
                <button type="submit" class="login-btn">Login</button>
            </form>
        </div>
        
        <div class="login-right">
            <div class="right-content">
                <h2>Welcome</h2>
                <p>Welcome to Warranty Management System</p>
                <p style="margin-top: 30px; font-size: 14px; opacity: 0.8;"></p>
            </div>
        </div>
    </div>
    
    <script>
        let currentCaptcha = '';
        
        async function loadCaptcha() {
            try {
                const response = await fetch('/api/captcha');
                const data = await response.json();
                currentCaptcha = data.captcha;
                document.getElementById('captchaImage').src = data.image;
            } catch (error) {
                console.error('Error loading CAPTCHA:', error);
            }
        }
        
        async function handleLogin(event) {
            event.preventDefault();
            
            const userId = document.getElementById('userId').value;
            const password = document.getElementById('password').value;
            const captchaInput = document.getElementById('captchaInput').value;
            const errorDiv = document.getElementById('errorMessage');
            
            if (captchaInput.toUpperCase() !== currentCaptcha) {
                errorDiv.textContent = ' CAPTCHA is incorrect';
                errorDiv.classList.add('show');
                loadCaptcha();
                return;
            }
            
            try {
                const response = await fetch('/api/login', {
                    method: 'POST',
                    headers: {
                        'Content-Type': 'application/json'
                    },
                    body: JSON.stringify({
                        user_id: userId,
                        password: password
                    })
                });
                
                if (response.ok) {
                    const data = await response.json();
                    window.location.href = '/dashboard';
                } else {
                    const error = await response.json();
                    errorDiv.textContent = ' ' + (error.detail || 'Login failed');
                    errorDiv.classList.add('show');
                    loadCaptcha();
                }
            } catch (error) {
                errorDiv.textContent = ' Error: ' + error.message;
                errorDiv.classList.add('show');
            }
        }
        
        window.onload = function() {
            loadCaptcha();
        };
    </script>
</body>
</html>
"""

# ==================== DASHBOARD HTML ====================

DASHBOARD_HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Unnati Warranty Management Dashboard</title>
    <style>
        * {
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }
        
        body {
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: linear-gradient(135deg, #f5f5f5 0%, #e0e0e0 100%);
            padding: 0;
            margin: 0;
            min-height: 100vh;
        }
        
        .navbar {
            background: linear-gradient(135deg, #FF8C00 0%, #FF6B35 100%);
            color: white;
            padding: 15px 0;
            box-shadow: 0 2px 8px rgba(0, 0, 0, 0.15);
            position: sticky;
            top: 0;
            z-index: 100;
        }
        
        .navbar .container-fluid {
            max-width: 1400px;
            margin: 0 auto;
            display: flex;
            justify-content: center;
            align-items: center;
            padding: 0 30px;
        }
        
        .navbar-brand {
            font-size: 24px;
            font-weight: 700;
        }
        
        .container {
            max-width: 1400px;
            margin: 30px auto;
            padding: 0 20px;
        }
        
        h1 {
            color: #333;
            margin-bottom: 30px;
            text-align: center;
            font-weight: 700;
        }
        
        .dashboard-content {
            background: white;
            border-radius: 12px;
            box-shadow: 0 2px 10px rgba(0, 0, 0, 0.1);
            padding: 30px;
        }
        
        .nav-tabs {
            border-bottom: 2px solid #FF8C00;
            margin-bottom: 30px;
        }
        
        .nav-tabs .nav-link {
            color: #666;
            font-weight: 600;
            border: none;
            border-bottom: 3px solid transparent;
            padding: 12px 20px;
            cursor: pointer;
            transition: all 0.3s ease;
        }
        
        .nav-tabs .nav-link:hover {
            color: #FF8C00;
            border-bottom-color: #FF8C00;
        }
        
        .nav-tabs .nav-link.active {
            color: #FF8C00;
            border-bottom-color: #FF8C00;
            background: transparent;
        }
        
        .tab-content {
            display: none;
        }
        
        .tab-content.active {
            display: block;
        }
        
        .data-table {
            width: 100%;
            border-collapse: collapse;
            margin-top: 20px;
            font-size: 12px;
            overflow-x: auto;
        }
        
        .data-table thead th {
            background: linear-gradient(135deg, #FF8C00 0%, #FF6B35 100%);
            color: white;
            padding: 12px;
            text-align: center;
            font-weight: 600;
            border: none;
            font-size: 11px;
        }
        
        .data-table tbody td {
            padding: 10px 12px;
            border-bottom: 1px solid #e0e0e0;
            text-align: right;
        }
        
        .data-table tbody td:first-child {
            text-align: left;
            font-weight: 600;
            color: #333;
        }
        
        .data-table tbody tr:hover {
            background: #f9f9f9;
        }
        
        .data-table tbody tr:last-child {
            background: #fff8f3;
            font-weight: 700;
            border-top: 2px solid #FF8C00;
            border-bottom: 2px solid #FF8C00;
        }
        
        .data-table tbody tr:last-child td {
            color: #FF8C00;
        }
        
        .loading-spinner {
            display: none;
            text-align: center;
            padding: 40px;
        }
        
        .spinner {
            border: 4px solid rgba(255, 140, 0, 0.2);
            border-top: 4px solid #FF8C00;
            border-radius: 50%;
            width: 40px;
            height: 40px;
            animation: spin 1s linear infinite;
            margin: 0 auto;
        }
        
        @keyframes spin {
            0% { transform: rotate(0deg); }
            100% { transform: rotate(360deg); }
        }
        
        .table-title {
            font-size: 16px;
            font-weight: 700;
            color: #FF8C00;
            margin-bottom: 15px;
        }
        
        .table-wrapper {
            overflow-x: auto;
        }
        
        .modal {
            display: none;
            position: fixed;
            z-index: 1000;
            left: 0;
            top: 0;
            width: 100%;
            height: 100%;
            background-color: rgba(0, 0, 0, 0.5);
            animation: fadeIn 0.3s ease;
        }
        
        .modal.show {
            display: flex;
            align-items: center;
            justify-content: center;
        }
        
        @keyframes fadeIn {
            from { opacity: 0; }
            to { opacity: 1; }
        }
        
        .modal-content {
            background: white;
            border-radius: 12px;
            box-shadow: 0 4px 20px rgba(0, 0, 0, 0.3);
            padding: 30px;
            width: 90%;
            max-width: 400px;
            animation: slideUp 0.3s ease;
        }
        
        @keyframes slideUp {
            from {
                transform: translateY(20px);
                opacity: 0;
            }
            to {
                transform: translateY(0);
                opacity: 1;
            }
        }
        
        .modal-header {
            font-size: 20px;
            font-weight: 700;
            color: #FF8C00;
            margin-bottom: 20px;
            display: flex;
            justify-content: space-between;
            align-items: center;
        }
        
        .modal-close-btn {
            background: none;
            border: none;
            font-size: 24px;
            cursor: pointer;
            color: #999;
            transition: color 0.3s ease;
        }
        
        .modal-close-btn:hover {
            color: #FF8C00;
        }
        
        .form-group {
            margin-bottom: 15px;
        }
        
        .form-group label {
            display: block;
            margin-bottom: 5px;
            color: #333;
            font-weight: 600;
            font-size: 13px;
        }
        
        .form-group input, .form-group select {
            width: 100%;
            padding: 10px;
            border: 2px solid #e0e0e0;
            border-radius: 6px;
            font-size: 13px;
            transition: all 0.3s ease;
        }
        
        .form-group input:focus, .form-group select:focus {
            outline: none;
            border-color: #FF8C00;
            box-shadow: 0 0 8px rgba(255, 140, 0, 0.2);
        }
        
        .modal-buttons {
            display: flex;
            gap: 10px;
            margin-top: 25px;
        }
        
        .modal-buttons button {
            flex: 1;
            padding: 10px;
            border: none;
            border-radius: 6px;
            font-weight: 600;
            cursor: pointer;
            transition: all 0.3s ease;
            font-size: 14px;
        }
        
        .btn-change {
            background: linear-gradient(135deg, #FF8C00 0%, #FF6B35 100%);
            color: white;
        }
        
        .btn-change:hover {
            transform: translateY(-2px);
            box-shadow: 0 5px 15px rgba(255, 140, 0, 0.3);
        }
        
        .btn-cancel {
            background: #e0e0e0;
            color: #333;
        }
        
        .btn-cancel:hover {
            background: #d0d0d0;
        }
        
        .modal-message {
            padding: 10px;
            border-radius: 6px;
            margin-bottom: 15px;
            font-size: 13px;
            display: none;
        }
        
        .modal-message.success {
            background: #e8f5e9;
            color: #2e7d32;
            border-left: 4px solid #2e7d32;
            display: block;
        }
        
        .modal-message.error {
            background: #ffebee;
            color: #c62828;
            border-left: 4px solid #c62828;
            display: block;
        }

        /* ===== EXPORT SECTION STYLES ===== */
        .export-section {
            margin: 30px 0;
            padding: 20px;
            background: linear-gradient(135deg, #fff8f3 0%, #ffe8d6 100%);
            border-radius: 8px;
            border-left: 5px solid #FF8C00;
            box-shadow: 0 2px 8px rgba(255, 140, 0, 0.1);
        }
        
        .export-section h3 {
            color: #FF8C00;
            margin-bottom: 15px;
            font-size: 16px;
            font-weight: 700;
            display: flex;
            align-items: center;
            gap: 8px;
        }
        
        .export-controls {
            display: flex;
            gap: 15px;
            align-items: center;
            flex-wrap: wrap;
            background: white;
            padding: 15px;
            border-radius: 6px;
        }
        
        .export-control-group {
            display: flex;
            gap: 8px;
            align-items: center;
        }
        
        .export-control-group label {
            font-weight: 600;
            color: #333;
            font-size: 14px;
            min-width: 80px;
        }
        
        .export-control-group select {
            padding: 8px 12px;
            border: 2px solid #FF8C00;
            border-radius: 4px;
            cursor: pointer;
            background: white;
            font-size: 13px;
            transition: all 0.3s ease;
            min-width: 150px;
        }
        
        .export-control-group select:hover {
            box-shadow: 0 2px 8px rgba(255, 140, 0, 0.2);
        }
        
        .export-control-group select:focus {
            outline: none;
            box-shadow: 0 0 8px rgba(255, 140, 0, 0.3);
        }
        
        .export-btn {
            padding: 10px 25px;
            background: linear-gradient(135deg, #4CAF50 0%, #45a049 100%);
            color: white;
            border: none;
            border-radius: 4px;
            cursor: pointer;
            font-weight: 700;
            font-size: 14px;
            transition: all 0.3s ease;
            display: flex;
            align-items: center;
            gap: 8px;
        }
        
        .export-btn:hover {
            background: linear-gradient(135deg, #45a049 0%, #3d8b40 100%);
            transform: translateY(-2px);
            box-shadow: 0 5px 15px rgba(76, 175, 80, 0.3);
        }
        
        .export-btn:active {
            transform: translateY(0);
        }

        .export-btn:disabled {
            background: #ccc;
            cursor: not-allowed;
            transform: none;
        }
    </style>
</head>
<body>
    <nav class="navbar navbar-dark">
        <div class="container-fluid">
            <span class="navbar-brand">Unnati Motors Warranty Management Dashboard</span>
        </div>
    </nav>
    
    <div class="container">
        
        <div class="dashboard-content">
            <div class="loading-spinner" id="loadingSpinner">
                <div class="spinner"></div>
                <p style="margin-top: 15px; color: #666;">Loading warranty data...</p>
            </div>
            
            <div id="warrantyTabs" style="display: none;">
                <!-- Tab Navigation -->
                <div class="nav-tabs">
                    <button class="nav-link active" onclick="switchTab('credit')"> Warranty Credit</button>
                    <button class="nav-link" onclick="switchTab('debit')"> Warranty Debit</button>
                    <button class="nav-link" onclick="switchTab('arbitration')"> Claim Arbitration</button>
                    <button class="nav-link" onclick="switchTab('currentmonth')"> Current Month Warranty</button>
                    <button class="nav-link" onclick="switchTab('compensation')"> Compensation Claim</button>
                    <button class="nav-link" onclick="switchTab('pr_approval')"> PR Approval</button>
                </div>

                <!-- EXPORT SECTION -->
                <div class="export-section">
                    <h3> Export to Excel</h3>
                    <div class="export-controls">
                        <div class="export-control-group">
                            <label for="divisionFilter">Division:</label>
                            <select id="divisionFilter">
                                <option value="">-- Select Division --</option>
                                <option value="All">All Divisions</option>
                            </select>
                        </div>
                        
                        <div class="export-control-group">
                            <label for="exportType">Export Type:</label>
                            <select id="exportType">
                                <option value="credit">Credit Note</option>
                                <option value="debit">Debit Note</option>
                                <option value="arbitration">Claim Arbitration</option>
                                <option value="currentmonth">Current Month Warranty</option>
                                <option value="compensation">Compensation Claim</option>
                                <option value="pr_approval">PR Approval</option>
                            </select>
                        </div>
                        
                        <button onclick="exportToExcel()" class="export-btn" id="exportBtn">
                             Export to Excel
                        </button>
                    </div>
                </div>
                
                <!-- Credit Note Tab -->
                <div id="credit" class="tab-content active">
                    <div class="table-title">Warranty Credit Note by Division & Month</div>
                    <div class="table-wrapper">
                        <table class="data-table" id="creditTable">
                            <thead></thead>
                            <tbody></tbody>
                        </table>
                    </div>
                </div>
                
                <!-- Debit Note Tab -->
                <div id="debit" class="tab-content">
                    <div class="table-title">Warranty Debit Note by Division & Month</div>
                    <div class="table-wrapper">
                        <table class="data-table" id="debitTable">
                            <thead></thead>
                            <tbody></tbody>
                        </table>
                    </div>
                </div>
                
                <!-- Claim Arbitration Tab -->
                <div id="arbitration" class="tab-content">
                    <div class="table-title">Warranty Claim Arbitration by Division</div>
                    <div class="table-wrapper">
                        <table class="data-table" id="arbitrationTable">
                            <thead></thead>
                            <tbody></tbody>
                        </table>
                    </div>
                </div>
                
                <!-- Current Month Warranty Tab -->
                <div id="currentmonth" class="tab-content">
                    <div class="table-title">Current Month Warranty - Pending Claims Summary</div>
                    <div class="table-wrapper">
                        <table class="data-table" id="currentMonthTable">
                            <thead></thead>
                            <tbody></tbody>
                        </table>
                    </div>
                </div>
                
                <!-- Compensation Claim Tab -->
                <div id="compensation" class="tab-content">
                    <div class="table-title">Compensation Claim - Transit Claims Summary</div>
                    <div class="table-wrapper">
                        <table class="data-table" id="compensationTable">
                            <thead></thead>
                            <tbody></tbody>
                        </table>
                    </div>
                </div>
                
                <!-- PR Approval Tab -->
                <div id="pr_approval" class="tab-content">
                    <div class="table-title">PR Approval - Claims Summary</div>
                    <div class="table-wrapper">
                        <table class="data-table" id="prApprovalTable">
                            <thead></thead>
                            <tbody></tbody>
                        </table>
                    </div>
                </div>
            </div>
        </div>
    </div>
    
    
    <script>
        let warrantyData = {};
        
        async function loadDashboard() {
            const spinner = document.getElementById('loadingSpinner');
            const tabs = document.getElementById('warrantyTabs');
            
            spinner.style.display = 'block';
            tabs.style.display = 'none';
            
            try {
                console.log('========== DASHBOARD LOAD START ==========');
                console.log(' Fetching warranty data with credentials...');
                console.log(' Current cookies:', document.cookie);
                
                const response = await fetch('/api/warranty-data', {
                    method: 'GET',
                    credentials: 'include',
                    headers: {
                        'Content-Type': 'application/json',
                        'Accept': 'application/json'
                    }
                });
                
                console.log(' Response received');
                console.log(' Response status:', response.status);
                
                if (response.status === 401) {
                    console.error(' Unauthorized (401) - Session expired');
                    alert('Session expired. Please login again.');
                    window.location.href = '/login-page';
                    return;
                }
                
                if (!response.ok) {
                    const text = await response.text();
                    console.error(' Response not OK:', response.status);
                    throw new Error('Failed to load warranty data: HTTP ' + response.status);
                }
                
                warrantyData = await response.json();
                console.log(' Warranty data loaded successfully');
                
                displayCreditTable(warrantyData.credit);
                displayDebitTable(warrantyData.debit);
                displayArbitrationTable(warrantyData.arbitration);
                displayCurrentMonthTable(warrantyData.currentMonth);
                displayCompensationTable(warrantyData.compensation);
                displayPrApprovalTable(warrantyData.prApproval);
                
                loadDivisions();
                
                spinner.style.display = 'none';
                tabs.style.display = 'block';
                console.log(' Dashboard rendered successfully');
            } catch (error) {
                console.error(' Error loading dashboard:', error);
                spinner.innerHTML = '<p style="color: red; padding: 20px; text-align: center;"> Error loading warranty data<br><br><button onclick="location.reload();" style="padding: 10px 20px; background: #FF8C00; color: white; border: none; border-radius: 6px; cursor: pointer; font-weight: 600;"> Refresh</button></p>';
            }
        }
        
        function displayCreditTable(data) {
            if (!data || data.length === 0) return;
            
            const table = document.getElementById('creditTable');
            const headers = Object.keys(data[0]);
            
            const headerRow = table.querySelector('thead');
            headerRow.innerHTML = headers.map(h => '<th>' + h + '</th>').join('');
            
            const tbody = table.querySelector('tbody');
            tbody.innerHTML = data.map((row) => {
                return '<tr>' + headers.map((h) => {
                    const value = typeof row[h] === 'number' ? row[h].toLocaleString('en-IN', {maximumFractionDigits: 0}) : row[h];
                    return '<td>' + value + '</td>';
                }).join('') + '</tr>';
            }).join('');
        }
        
        function displayDebitTable(data) {
            if (!data || data.length === 0) return;
            
            const table = document.getElementById('debitTable');
            const headers = Object.keys(data[0]);
            
            const headerRow = table.querySelector('thead');
            headerRow.innerHTML = headers.map(h => '<th>' + h + '</th>').join('');
            
            const tbody = table.querySelector('tbody');
            tbody.innerHTML = data.map((row) => {
                return '<tr>' + headers.map((h) => {
                    const value = typeof row[h] === 'number' ? row[h].toLocaleString('en-IN', {maximumFractionDigits: 0}) : row[h];
                    return '<td>' + value + '</td>';
                }).join('') + '</tr>';
            }).join('');
        }
        
        function displayArbitrationTable(data) {
            if (!data || data.length === 0) return;
            
            const table = document.getElementById('arbitrationTable');
            const headers = Object.keys(data[0]);
            
            const headerRow = table.querySelector('thead');
            headerRow.innerHTML = headers.map(h => '<th>' + h + '</th>').join('');
            
            const tbody = table.querySelector('tbody');
            tbody.innerHTML = data.map((row) => {
                return '<tr>' + headers.map((h) => {
                    const value = typeof row[h] === 'number' ? row[h].toLocaleString('en-IN', {maximumFractionDigits: 0}) : row[h];
                    return '<td>' + value + '</td>';
                }).join('') + '</tr>';
            }).join('');
        }
        
        function displayCurrentMonthTable(data) {
            if (!data || data.length === 0) return;
            
            const table = document.getElementById('currentMonthTable');
            const headers = Object.keys(data[0]);
            
            const headerRow = table.querySelector('thead');
            headerRow.innerHTML = headers.map(h => '<th>' + h + '</th>').join('');
            
            const tbody = table.querySelector('tbody');
            tbody.innerHTML = data.map((row) => {
                return '<tr>' + headers.map((h) => {
                    const value = typeof row[h] === 'number' ? row[h].toLocaleString('en-IN', {maximumFractionDigits: 0}) : row[h];
                    return '<td>' + value + '</td>';
                }).join('') + '</tr>';
            }).join('');
        }
        
        function displayCompensationTable(data) {
            if (!data || data.length === 0) return;
            
            const table = document.getElementById('compensationTable');
            const headers = Object.keys(data[0]);
            
            const headerRow = table.querySelector('thead');
            headerRow.innerHTML = headers.map(h => '<th>' + h + '</th>').join('');
            
            const tbody = table.querySelector('tbody');
            tbody.innerHTML = data.map((row) => {
                return '<tr>' + headers.map((h) => {
                    const value = typeof row[h] === 'number' ? row[h].toLocaleString('en-IN', {maximumFractionDigits: 2}) : row[h];
                    return '<td>' + value + '</td>';
                }).join('') + '</tr>';
            }).join('');
        }
        
        function displayPrApprovalTable(data) {
            if (!data || data.length === 0) return;
            
            const table = document.getElementById('prApprovalTable');
            const headers = Object.keys(data[0]);
            
            const headerRow = table.querySelector('thead');
            headerRow.innerHTML = headers.map(h => '<th>' + h + '</th>').join('');
            
            const tbody = table.querySelector('tbody');
            tbody.innerHTML = data.map((row) => {
                return '<tr>' + headers.map((h) => {
                    const value = typeof row[h] === 'number' ? row[h].toLocaleString('en-IN', {maximumFractionDigits: 2}) : row[h];
                    return '<td>' + value + '</td>';
                }).join('') + '</tr>';
            }).join('');
        }
        
        function switchTab(tabName) {
            document.querySelectorAll('.tab-content').forEach(tab => {
                tab.classList.remove('active');
            });
            
            document.querySelectorAll('.nav-link').forEach(btn => {
                btn.classList.remove('active');
            });
            
            document.getElementById(tabName).classList.add('active');
            event.target.classList.add('active');
        }

        // ===== EXPORT FUNCTIONS =====
        function loadDivisions() {
            console.log(' Loading divisions from warranty data...');
            const divisions = new Set();
            
            const currentType = document.getElementById('exportType').value;
            let dataSource = warrantyData.credit;
            
            if (currentType === 'debit') dataSource = warrantyData.debit;
            if (currentType === 'arbitration') dataSource = warrantyData.arbitration;
            if (currentType === 'currentmonth') dataSource = warrantyData.currentMonth;
            if (currentType === 'compensation') dataSource = warrantyData.compensation;
            if (currentType === 'pr_approval') dataSource = warrantyData.prApproval;
            
            if (dataSource && dataSource.length > 0) {
                dataSource.forEach(row => {
                    if (row.Division && row.Division !== 'Grand Total') {
                        divisions.add(row.Division);
                    }
                });
            }
            
            const divisionSelect = document.getElementById('divisionFilter');
            const currentValue = divisionSelect.value;
            
            divisionSelect.innerHTML = '<option value="">-- Select Division --</option><option value="All">All Divisions</option>';
            
            Array.from(divisions).sort().forEach(div => {
                const option = document.createElement('option');
                option.value = div;
                option.textContent = div;
                divisionSelect.appendChild(option);
            });
            
            if (currentValue && divisionSelect.querySelector(`option[value="${currentValue}"]`)) {
                divisionSelect.value = currentValue;
            }
            
            console.log(' Divisions loaded:', Array.from(divisions).length);
        }

        // Listen for export type changes
        document.getElementById('exportType')?.addEventListener('change', loadDivisions);

        async function exportToExcel() {
            const division = document.getElementById('divisionFilter').value;
            const type = document.getElementById('exportType').value;
            const exportBtn = document.getElementById('exportBtn');
            
            if (!division) {
                alert(' Please select a division');
                return;
            }
            
            console.log(` Exporting ${type} data for division: ${division}`);
            exportBtn.disabled = true;
            exportBtn.textContent = '⏳ Exporting...';
            
            try {
                const response = await fetch('/api/export-to-excel', {
                    method: 'POST',
                    headers: {
                        'Content-Type': 'application/json'
                    },
                    credentials: 'include',
                    body: JSON.stringify({
                        division: division,
                        type: type
                    })
                });
                
                if (!response.ok) {
                    const error = await response.json().catch(() => ({detail: 'Export failed'}));
                    throw new Error(error.detail || 'Export failed');
                }
                
                const blob = await response.blob();
                const url = window.URL.createObjectURL(blob);
                const a = document.createElement('a');
                a.href = url;
                a.download = `${type}_${division}_${new Date().toISOString().split('T')[0]}.xlsx`;
                document.body.appendChild(a);
                a.click();
                window.URL.revokeObjectURL(url);
                document.body.removeChild(a);
                
                console.log(' Export completed successfully');
                alert(' Export completed successfully!');
            } catch (error) {
                console.error(' Export error:', error);
                alert(' Export failed: ' + error.message);
            } finally {
                exportBtn.disabled = false;
                exportBtn.textContent = ' Export to Excel';
            }
        }
        
        window.onload = function() {
            console.log('========== DASHBOARD PAGE ONLOAD ==========');
            console.log(' Dashboard page loaded');
            
            loadDashboard();
        };
    </script>
</body>
</html>
"""

# ==================== FASTAPI SETUP ====================

app = FastAPI()

# ==================== API ENDPOINTS ====================

@app.post("/api/change-password")
async def change_password(request: Request, session_id: str = Cookie(None)):
    """Change user password"""
    try:
        if not session_id or not verify_session(session_id):
            raise HTTPException(status_code=401, detail="Not authenticated")
        
        user_id = verify_session(session_id)
        body = await request.json()
        current_password = body.get('current_password', '')
        new_password = body.get('new_password', '')
        
        if not current_password or not new_password:
            raise HTTPException(status_code=400, detail="Missing required fields")
        
        if current_password != USER_CREDENTIALS.get(user_id):
            raise HTTPException(status_code=401, detail="Current password is incorrect")
        
        if len(new_password) < 6:
            raise HTTPException(status_code=400, detail="New password must be at least 6 characters")
        
        success = update_user_password_in_excel(user_id, new_password)
        
        if not success:
            raise HTTPException(status_code=500, detail="Failed to update password in database")
        
        USER_CREDENTIALS[user_id] = new_password
        
        print(f" Password successfully changed for user {user_id}")
        
        return {
            "success": True,
            "message": "Password changed successfully! You can now login with your new password."
        }
        
    except HTTPException as e:
        raise
    except Exception as e:
        print(f" Error changing password: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/export-to-excel")
async def export_to_excel(request: Request):
    """Export selected division data to Excel with summary and detailed sheets"""
    try:
        print(f" Export request received")
        
        # Get request body
        body = await request.json()
        selected_division = body.get('division', 'All')
        export_type = body.get('type', 'credit')
        
        print(f" Export Type: {export_type}, Division: {selected_division}")
        
        # Validate export type
        if export_type not in ['credit', 'debit', 'arbitration', 'currentmonth', 'compensation', 'pr_approval']:
            raise HTTPException(status_code=400, detail="Invalid export type")
        
        # Handle Current Month Warranty export separately
        if export_type == 'currentmonth':
            return await export_current_month_warranty(selected_division)
        
        # Handle Compensation Claim export separately
        if export_type == 'compensation':
            return await export_compensation_claim(selected_division)
        
        # Handle PR Approval export separately
        if export_type == 'pr_approval':
            return await export_pr_approval(selected_division)
        
        # Get the appropriate dataframe
        if export_type == 'credit':
            df = WARRANTY_DATA['credit_df']
        elif export_type == 'debit':
            df = WARRANTY_DATA['debit_df']
        else:  # arbitration
            df = WARRANTY_DATA['arbitration_df']
        
        if df is None or df.empty:
            raise HTTPException(status_code=500, detail="No data available for export")
        
        print(f" Original data rows: {len(df)}")
        
        # Reverse dealer mapping
        dealer_mapping = {
            'AMRAVATI': 'AMT',
            'CHAUFULA_SZZ': 'CHA',
            'CHIKHALI': 'CHI',
            'KOLHAPUR_WS': 'KOL',
            'NAGPUR_KAMPTHEE ROAD': 'HO',
            'NAGPUR_WARDHAMAN NGR': 'CITY',
            'SHIKRAPUR_SZS': 'SHI',
            'WAGHOLI': 'WAG',
            'YAVATMAL': 'YAT',
            'NAGPUR_WARDHAMAN NGR_CQ': 'CQ'
        }
        reverse_mapping = {v: k for k, v in dealer_mapping.items()}
        
        # Filter by division if not "All"
        if selected_division != 'All' and selected_division != 'Grand Total':
            df_export = df[df['Division'] == selected_division].copy()
            # Add Grand Total row if exists
            grand_total_row = df[df['Division'] == 'Grand Total']
            if not grand_total_row.empty:
                df_export = pd.concat([df_export, grand_total_row], ignore_index=True)
        else:
            df_export = df.copy()
        
        print(f" Filtered data rows: {len(df_export)}")
        
        # Create workbook with styling
        wb = Workbook()
        
        # Define styles
        header_fill = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # ==================== SHEET 1: SUMMARY ====================
        ws1 = wb.active
        if selected_division != 'All' and selected_division != 'Grand Total':
            ws1.title = f"{selected_division} - {export_type.capitalize()}"
        else:
            ws1.title = export_type.capitalize()
        
        # Write headers for sheet 1
        for col_idx, column in enumerate(df_export.columns, 1):
            cell = ws1.cell(row=1, column=col_idx, value=column)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Write data for sheet 1
        for row_idx, row in enumerate(df_export.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws1.cell(row=row_idx, column=col_idx)
                
                # Format the value
                if isinstance(value, (int, float)):
                    cell.value = value
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                else:
                    cell.value = str(value)
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                
                cell.border = border
        
        # Adjust column widths for sheet 1
        for col_idx, column in enumerate(df_export.columns, 1):
            max_length = min(max(
                df_export[column].astype(str).map(len).max(),
                len(str(column))
            ) + 2, 30)
            ws1.column_dimensions[get_column_letter(col_idx)].width = max_length
        
        # ==================== SHEET 2: DETAILED SOURCE DATA ====================
        if selected_division != 'All' and selected_division != 'Grand Total':
            ws2 = wb.create_sheet()
            ws2.title = f"{selected_division} - Detailed Data"
            
            # Get the dealer location for the selected division
            dealer_location = reverse_mapping.get(selected_division)
            
            if dealer_location and WARRANTY_DATA['source_df'] is not None:
                source_df = WARRANTY_DATA['source_df'].copy()
                
                # Filter by dealer location
                detail_df = source_df[source_df['Dealer Location'] == dealer_location].copy()
                
                # Define all required columns
                required_columns = [
                    'Fiscal Month',
                    'Dealer Location',
                    'Claim arbitration ID',
                    'Claim Invoice Date',
                    'Claim No',
                    'Claim Date',
                    'Chassis No',
                    'Ro Id',
                    'Claim Type'
                ]
                
                # Add amount columns based on export type
                if export_type == 'arbitration':
                    required_columns.append('Credit Note Amount')
                else:
                    required_columns.append('Total Claim Amount')
                
                # Helper function to check if Claim arbitration ID is empty or contains "-"
                def is_empty_or_hyphen(value):
                    if pd.isna(value): 
                        return True
                    value = str(value).strip()
                    if value == '' or value == '-' or value.upper() == 'NAN':
                        return True
                    return False
                
                # Helper function to check if arbitration ID has valid ARB number
                def has_valid_arb_id(value):
                    if pd.isna(value): 
                        return False
                    value = str(value).strip().upper()
                    return value.startswith('ARB') and value != 'NAN' and value != ''
                
                # Further filter by export type and add type-specific columns
                if export_type == 'credit':
                    detail_df = detail_df[detail_df['Credit Note Amount'] > 0].copy()
                    detail_df = detail_df[detail_df['Claim arbitration ID'].apply(is_empty_or_hyphen)].copy()
                    required_columns.append('Credit Note Amount')
                    
                elif export_type == 'debit':
                    detail_df = detail_df[detail_df['Debit Note Amount'] > 0].copy()
                    required_columns.append('Debit Note Amount')
                    
                else:  # arbitration
                    detail_df = detail_df[detail_df['Claim arbitration ID'].apply(has_valid_arb_id)].copy()
                    required_columns.append('Debit Note Amount')
                
                # Select only the required columns that exist
                available_columns = [col for col in required_columns if col in detail_df.columns]
                detail_df = detail_df[available_columns].copy()
                
                # Format Claim No as text
                if 'Claim No' in detail_df.columns:
                    def format_claim_no(x):
                        if pd.isna(x) or str(x).strip() == '':
                            return ''
                        try:
                            return str(int(float(x)))
                        except (ValueError, TypeError):
                            return str(x).strip()
                    
                    detail_df['Claim No'] = detail_df['Claim No'].apply(format_claim_no)
                
                # Add "RO" prefix to Ro Id
                if 'Ro Id' in detail_df.columns:
                    def format_ro_id(x):
                        if pd.isna(x) or str(x).strip() == '':
                            return ''
                        try:
                            return f"RO{str(int(float(x)))}"
                        except (ValueError, TypeError):
                            value_str = str(x).strip()
                            if not value_str.startswith('RO'):
                                return f"RO{value_str}"
                            return value_str
                    
                    detail_df['Ro Id'] = detail_df['Ro Id'].apply(format_ro_id)
                
                # Rename the amount column for arbitration
                if export_type == 'arbitration' and 'Debit Note Amount' in detail_df.columns:
                    detail_df = detail_df.rename(columns={'Debit Note Amount': 'Arbitration Amount'})
                
                # Sort by Fiscal Month
                month_order = ['Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec', 'Jan', 'Feb', 'Mar']
                detail_df['Month'] = detail_df['Fiscal Month'].astype(str).str.strip().str[:3]
                detail_df['Month_Order'] = detail_df['Month'].apply(lambda x: month_order.index(x) if x in month_order else 999)
                detail_df = detail_df.sort_values('Month_Order').drop(['Month', 'Month_Order'], axis=1)
                
                print(f" Detailed data rows for {selected_division}: {len(detail_df)}")
                
                # Write headers for sheet 2
                for col_idx, column in enumerate(detail_df.columns, 1):
                    cell = ws2.cell(row=1, column=col_idx, value=column)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                # Write data for sheet 2
                for row_idx, row in enumerate(detail_df.itertuples(index=False), 2):
                    for col_idx, value in enumerate(row, 1):
                        cell = ws2.cell(row=row_idx, column=col_idx)
                        column_name = detail_df.columns[col_idx - 1]
                        
                        if column_name == 'Claim No':
                            cell.value = str(value) if not pd.isna(value) and str(value).strip() != '' else ''
                            cell.alignment = Alignment(horizontal='left', vertical='center')
                        elif column_name == 'Ro Id':
                            cell.value = str(value) if not pd.isna(value) and str(value).strip() != '' else ''
                            cell.alignment = Alignment(horizontal='left', vertical='center')
                        elif isinstance(value, (int, float)):
                            cell.value = value
                            cell.number_format = '#,##0.00'
                            cell.alignment = Alignment(horizontal='right', vertical='center')
                        elif isinstance(value, (datetime, pd.Timestamp)):
                            cell.value = value
                            cell.number_format = 'mm-dd-yyyy'
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        else:
                            cell.value = str(value) if not pd.isna(value) else ''
                            cell.alignment = Alignment(horizontal='left', vertical='center')
                        
                        cell.border = border
                
                # Adjust column widths for sheet 2
                for col_idx, column in enumerate(detail_df.columns, 1):
                    max_length = min(max(
                        detail_df[column].astype(str).map(len).max(),
                        len(str(column))
                    ) + 2, 30)
                    column_letter = ws2.cell(row=1, column=col_idx).column_letter
                    ws2.column_dimensions[column_letter].width = max_length
                
                # ==================== SHEET 3: PENDING ARBITRATION (Only for Arbitration Export) ====================
                if export_type == 'arbitration':
                    ws3 = wb.create_sheet()
                    ws3.title = f"{selected_division} - Pending Arb"
                    
                    # Get pending arbitration records
                    pending_df = source_df[source_df['Dealer Location'] == dealer_location].copy()
                    pending_df = pending_df[pending_df['Debit Note Amount'] > 0].copy()
                    pending_df = pending_df[pending_df['Claim arbitration ID'].apply(is_empty_or_hyphen)].copy()
                    
                    # Define columns for pending arbitration
                    pending_columns = [
                        'Fiscal Month',
                        'Dealer Location',
                        'Claim arbitration ID',
                        'Claim Invoice Date',
                        'Claim No',
                        'Claim Date',
                        'Chassis No',
                        'Ro Id',
                        'Claim Type',
                        'Credit Note Amount',
                        'Debit Note Amount'
                    ]
                    
                    # Select available columns
                    available_pending_columns = [col for col in pending_columns if col in pending_df.columns]
                    pending_df = pending_df[available_pending_columns].copy()
                    
                    # Format Claim No as text
                    if 'Claim No' in pending_df.columns:
                        def format_claim_no(x):
                            if pd.isna(x) or str(x).strip() == '':
                                return ''
                            try:
                                return str(int(float(x)))
                            except (ValueError, TypeError):
                                return str(x).strip()
                        
                        pending_df['Claim No'] = pending_df['Claim No'].apply(format_claim_no)
                    
                    # Add "RO" prefix to Ro Id
                    if 'Ro Id' in pending_df.columns:
                        def format_ro_id(x):
                            if pd.isna(x) or str(x).strip() == '':
                                return ''
                            try:
                                return f"RO{str(int(float(x)))}"
                            except (ValueError, TypeError):
                                value_str = str(x).strip()
                                if not value_str.startswith('RO'):
                                    return f"RO{value_str}"
                                return value_str
                        
                        pending_df['Ro Id'] = pending_df['Ro Id'].apply(format_ro_id)
                    
                    # Rename for clarity
                    if 'Debit Note Amount' in pending_df.columns:
                        pending_df = pending_df.rename(columns={'Debit Note Amount': 'Pending Arbitration Amount'})
                    
                    # Sort by Fiscal Month
                    pending_df['Month'] = pending_df['Fiscal Month'].astype(str).str.strip().str[:3]
                    pending_df['Month_Order'] = pending_df['Month'].apply(lambda x: month_order.index(x) if x in month_order else 999)
                    pending_df = pending_df.sort_values('Month_Order').drop(['Month', 'Month_Order'], axis=1)
                    
                    print(f" Pending Arbitration rows for {selected_division}: {len(pending_df)}")
                    
                    # Write headers for sheet 3
                    for col_idx, column in enumerate(pending_df.columns, 1):
                        cell = ws3.cell(row=1, column=col_idx, value=column)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.border = border
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    
                    # Write data for sheet 3
                    for row_idx, row in enumerate(pending_df.itertuples(index=False), 2):
                        for col_idx, value in enumerate(row, 1):
                            cell = ws3.cell(row=row_idx, column=col_idx)
                            column_name = pending_df.columns[col_idx - 1]
                            
                            if column_name == 'Claim No':
                                cell.value = str(value) if not pd.isna(value) and str(value).strip() != '' else ''
                                cell.alignment = Alignment(horizontal='left', vertical='center')
                            elif column_name == 'Ro Id':
                                cell.value = str(value) if not pd.isna(value) and str(value).strip() != '' else ''
                                cell.alignment = Alignment(horizontal='left', vertical='center')
                            elif isinstance(value, (int, float)):
                                cell.value = value
                                cell.number_format = '#,##0.00'
                                cell.alignment = Alignment(horizontal='right', vertical='center')
                            elif isinstance(value, (datetime, pd.Timestamp)):
                                cell.value = value
                                cell.number_format = 'mm-dd-yyyy'
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                            else:
                                cell.value = str(value) if not pd.isna(value) else ''
                                cell.alignment = Alignment(horizontal='left', vertical='center')
                            
                            cell.border = border
                    
                    # Adjust column widths for sheet 3
                    for col_idx, column in enumerate(pending_df.columns, 1):
                        max_length = min(max(
                            pending_df[column].astype(str).map(len).max(),
                            len(str(column))
                        ) + 2, 30)
                        column_letter = ws3.cell(row=1, column=col_idx).column_letter
                        ws3.column_dimensions[column_letter].width = max_length
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"{selected_division}_{export_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        print(f" Export file prepared: {filename}")
        
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except HTTPException as e:
        raise
    except Exception as e:
        print(f" Export error: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Export error: {str(e)}")

async def export_current_month_warranty(selected_division: str):
    """Export Current Month Warranty data"""
    try:
        summary_df = WARRANTY_DATA['current_month_df']
        source_df = WARRANTY_DATA['current_month_source_df']
        
        if summary_df is None or summary_df.empty:
            raise HTTPException(status_code=500, detail="No current month warranty data available")
        
        # Filter by division if not "All"
        if selected_division != 'All' and selected_division != 'Grand Total':
            df_export = summary_df[summary_df['Division'] == selected_division].copy()
            grand_total_row = summary_df[summary_df['Division'] == 'Grand Total']
            if not grand_total_row.empty:
                df_export = pd.concat([df_export, grand_total_row], ignore_index=True)
        else:
            df_export = summary_df.copy()
        
        # Create workbook
        wb = Workbook()
        
        # Define styles
        header_fill = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # ==================== SHEET 1: SUMMARY ====================
        ws1 = wb.active
        if selected_division != 'All' and selected_division != 'Grand Total':
            ws1.title = f"{selected_division} - Summary"
        else:
            ws1.title = "Current Month Summary"
        
        # Write headers
        for col_idx, column in enumerate(df_export.columns, 1):
            cell = ws1.cell(row=1, column=col_idx, value=column)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Write data
        for row_idx, row in enumerate(df_export.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws1.cell(row=row_idx, column=col_idx)
                
                if isinstance(value, (int, float)):
                    cell.value = value
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                else:
                    cell.value = str(value)
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                
                cell.border = border
        
        # Adjust column widths
        for col_idx, column in enumerate(df_export.columns, 1):
            max_length = min(max(
                df_export[column].astype(str).map(len).max(),
                len(str(column))
            ) + 2, 30)
            ws1.column_dimensions[get_column_letter(col_idx)].width = max_length
        
        # ==================== SHEET 2: PENDING SPARES CLAIMS ====================
        if source_df is not None and not source_df.empty:
            # Filter source data by division if specific division selected
            if selected_division != 'All' and selected_division != 'Grand Total':
                spares_df = source_df[source_df['Division'] == selected_division].copy()
            else:
                spares_df = source_df.copy()
            
            # Filter only records where Pending Claims Spares is NOT empty
            spares_df = spares_df[spares_df['Pending Claims Spares'].notna()].copy()
            
            if not spares_df.empty:
                ws2 = wb.create_sheet()
                if selected_division != 'All' and selected_division != 'Grand Total':
                    ws2.title = f"{selected_division} - Spares"
                else:
                    ws2.title = "Pending Spares Claims"
                
                # Write headers
                for col_idx, column in enumerate(spares_df.columns, 1):
                    cell = ws2.cell(row=1, column=col_idx, value=column)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                # Write data
                for row_idx, row in enumerate(spares_df.itertuples(index=False), 2):
                    for col_idx, value in enumerate(row, 1):
                        cell = ws2.cell(row=row_idx, column=col_idx)
                        
                        if isinstance(value, (int, float)):
                            cell.value = value
                            cell.number_format = '#,##0.00'
                            cell.alignment = Alignment(horizontal='right', vertical='center')
                        elif isinstance(value, (datetime, pd.Timestamp)):
                            cell.value = value
                            cell.number_format = 'mm-dd-yyyy'
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        else:
                            cell.value = str(value) if not pd.isna(value) else ''
                            cell.alignment = Alignment(horizontal='left', vertical='center')
                        
                        cell.border = border
                
                # Adjust column widths
                for col_idx, column in enumerate(spares_df.columns, 1):
                    max_length = min(max(
                        spares_df[column].astype(str).map(len).max(),
                        len(str(column))
                    ) + 2, 35)
                    column_letter = ws2.cell(row=1, column=col_idx).column_letter
                    ws2.column_dimensions[column_letter].width = max_length
                
                print(f" Pending Spares Claims rows: {len(spares_df)}")
        
        # ==================== SHEET 3: PENDING LABOUR CLAIMS ====================
        if source_df is not None and not source_df.empty:
            # Filter source data by division if specific division selected
            if selected_division != 'All' and selected_division != 'Grand Total':
                labour_df = source_df[source_df['Division'] == selected_division].copy()
            else:
                labour_df = source_df.copy()
            
            # Filter only records where Pending Claims Labour is NOT empty
            labour_df = labour_df[labour_df['Pending Claims Labour'].notna()].copy()
            
            if not labour_df.empty:
                ws3 = wb.create_sheet()
                if selected_division != 'All' and selected_division != 'Grand Total':
                    ws3.title = f"{selected_division} - Labour"
                else:
                    ws3.title = "Pending Labour Claims"
                
                # Write headers
                for col_idx, column in enumerate(labour_df.columns, 1):
                    cell = ws3.cell(row=1, column=col_idx, value=column)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                # Write data
                for row_idx, row in enumerate(labour_df.itertuples(index=False), 2):
                    for col_idx, value in enumerate(row, 1):
                        cell = ws3.cell(row=row_idx, column=col_idx)
                        
                        if isinstance(value, (int, float)):
                            cell.value = value
                            cell.number_format = '#,##0.00'
                            cell.alignment = Alignment(horizontal='right', vertical='center')
                        elif isinstance(value, (datetime, pd.Timestamp)):
                            cell.value = value
                            cell.number_format = 'mm-dd-yyyy'
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        else:
                            cell.value = str(value) if not pd.isna(value) else ''
                            cell.alignment = Alignment(horizontal='left', vertical='center')
                        
                        cell.border = border
                
                # Adjust column widths
                for col_idx, column in enumerate(labour_df.columns, 1):
                    max_length = min(max(
                        labour_df[column].astype(str).map(len).max(),
                        len(str(column))
                    ) + 2, 35)
                    column_letter = ws3.cell(row=1, column=col_idx).column_letter
                    ws3.column_dimensions[column_letter].width = max_length
                
                print(f" Pending Labour Claims rows: {len(labour_df)}")
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"{selected_division}_CurrentMonthWarranty_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        print(f" Current Month Warranty export completed: {filename}")
        
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        print(f" Current Month Warranty export error: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Export error: {str(e)}")

async def export_compensation_claim(selected_division: str):
    """Export Compensation Claim data"""
    try:
        summary_df = WARRANTY_DATA['compensation_df']
        source_df = WARRANTY_DATA['compensation_source_df']
        
        if summary_df is None or summary_df.empty:
            raise HTTPException(status_code=500, detail="No compensation claim data available")
        
        # Filter by division if not "All"
        if selected_division != 'All' and selected_division != 'Grand Total':
            df_export = summary_df[summary_df['Division'] == selected_division].copy()
            grand_total_row = summary_df[summary_df['Division'] == 'Grand Total']
            if not grand_total_row.empty:
                df_export = pd.concat([df_export, grand_total_row], ignore_index=True)
        else:
            df_export = summary_df.copy()
        
        # Create workbook
        wb = Workbook()
        
        # Define styles
        header_fill = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # ==================== SHEET 1: SUMMARY ====================
        ws1 = wb.active
        if selected_division != 'All' and selected_division != 'Grand Total':
            ws1.title = f"{selected_division} - Summary"
        else:
            ws1.title = "Compensation Summary"
        
        # Write headers
        for col_idx, column in enumerate(df_export.columns, 1):
            cell = ws1.cell(row=1, column=col_idx, value=column)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Write data
        for row_idx, row in enumerate(df_export.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws1.cell(row=row_idx, column=col_idx)
                
                if isinstance(value, (int, float)):
                    cell.value = value
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                else:
                    cell.value = str(value)
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                
                cell.border = border
        
        # Adjust column widths
        for col_idx, column in enumerate(df_export.columns, 1):
            max_length = min(max(
                df_export[column].astype(str).map(len).max(),
                len(str(column))
            ) + 2, 30)
            ws1.column_dimensions[get_column_letter(col_idx)].width = max_length
        
        # ==================== SHEET 2: DETAILED COMPENSATION CLAIMS ====================
        if source_df is not None and not source_df.empty:
            # Filter source data by division if specific division selected
            if selected_division != 'All' and selected_division != 'Grand Total':
                detail_df = source_df[source_df['Division'] == selected_division].copy()
            else:
                detail_df = source_df.copy()
            
            if not detail_df.empty:
                ws2 = wb.create_sheet()
                if selected_division != 'All' and selected_division != 'Grand Total':
                    ws2.title = f"{selected_division} - Details"
                else:
                    ws2.title = "Compensation Details"
                
                # Write headers
                for col_idx, column in enumerate(detail_df.columns, 1):
                    cell = ws2.cell(row=1, column=col_idx, value=column)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                # Write data
                for row_idx, row in enumerate(detail_df.itertuples(index=False), 2):
                    for col_idx, value in enumerate(row, 1):
                        cell = ws2.cell(row=row_idx, column=col_idx)
                        column_name = detail_df.columns[col_idx - 1]
                        
                        if column_name == 'RO Id.':
                            cell.value = str(value) if not pd.isna(value) and str(value).strip() != '' else ''
                            cell.alignment = Alignment(horizontal='left', vertical='center')
                        elif isinstance(value, (int, float)):
                            cell.value = value
                            cell.number_format = '#,##0.00'
                            cell.alignment = Alignment(horizontal='right', vertical='center')
                        elif isinstance(value, (datetime, pd.Timestamp)):
                            cell.value = value
                            cell.number_format = 'mm-dd-yyyy'
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        else:
                            cell.value = str(value) if not pd.isna(value) else ''
                            cell.alignment = Alignment(horizontal='left', vertical='center')
                        
                        cell.border = border
                
                # Adjust column widths
                for col_idx, column in enumerate(detail_df.columns, 1):
                    max_length = min(max(
                        detail_df[column].astype(str).map(len).max(),
                        len(str(column))
                    ) + 2, 35)
                    column_letter = ws2.cell(row=1, column=col_idx).column_letter
                    ws2.column_dimensions[column_letter].width = max_length
                
                print(f" Compensation Claim details rows: {len(detail_df)}")
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"{selected_division}_CompensationClaim_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        print(f" Compensation Claim export completed: {filename}")
        
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        print(f" Compensation Claim export error: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Export error: {str(e)}")

async def export_pr_approval(selected_division: str):
    """Export PR Approval data"""
    try:
        summary_df = WARRANTY_DATA['pr_approval_df']
        source_df = WARRANTY_DATA['pr_approval_source_df']
        
        if summary_df is None or summary_df.empty:
            raise HTTPException(status_code=500, detail="No PR Approval data available")
        
        # Filter by division if not "All"
        if selected_division != 'All' and selected_division != 'Grand Total':
            df_export = summary_df[summary_df['Division'] == selected_division].copy()
            grand_total_row = summary_df[summary_df['Division'] == 'Grand Total']
            if not grand_total_row.empty:
                df_export = pd.concat([df_export, grand_total_row], ignore_index=True)
        else:
            df_export = summary_df.copy()
        
        # Create workbook
        wb = Workbook()
        
        # Define styles
        header_fill = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # ==================== SHEET 1: SUMMARY ====================
        ws1 = wb.active
        if selected_division != 'All' and selected_division != 'Grand Total':
            ws1.title = f"{selected_division} - Summary"
        else:
            ws1.title = "PR Approval Summary"
        
        # Write headers
        for col_idx, column in enumerate(df_export.columns, 1):
            cell = ws1.cell(row=1, column=col_idx, value=column)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Write data
        for row_idx, row in enumerate(df_export.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws1.cell(row=row_idx, column=col_idx)
                
                if isinstance(value, (int, float)):
                    cell.value = value
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                else:
                    cell.value = str(value)
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                
                cell.border = border
        
        # Adjust column widths
        for col_idx, column in enumerate(df_export.columns, 1):
            max_length = min(max(
                df_export[column].astype(str).map(len).max(),
                len(str(column))
            ) + 2, 30)
            ws1.column_dimensions[get_column_letter(col_idx)].width = max_length
        
        # ==================== SHEET 2: COMPLETE DETAILED DATA ====================
        if source_df is not None and not source_df.empty:
            # Filter source data by division if specific division selected
            if selected_division != 'All' and selected_division != 'Grand Total':
                detail_df = source_df[source_df['Division'] == selected_division].copy()
            else:
                detail_df = source_df.copy()
            
            if not detail_df.empty:
                ws2 = wb.create_sheet()
                if selected_division != 'All' and selected_division != 'Grand Total':
                    ws2.title = f"{selected_division} - Details"
                else:
                    ws2.title = "PR Approval Details"
                
                # Write headers
                for col_idx, column in enumerate(detail_df.columns, 1):
                    cell = ws2.cell(row=1, column=col_idx, value=column)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                # Write data
                for row_idx, row in enumerate(detail_df.itertuples(index=False), 2):
                    for col_idx, value in enumerate(row, 1):
                        cell = ws2.cell(row=row_idx, column=col_idx)
                        
                        if isinstance(value, (int, float)):
                            cell.value = value
                            cell.number_format = '#,##0.00'
                            cell.alignment = Alignment(horizontal='right', vertical='center')
                        elif isinstance(value, (datetime, pd.Timestamp)):
                            cell.value = value
                            cell.number_format = 'mm-dd-yyyy'
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        else:
                            cell.value = str(value) if not pd.isna(value) else ''
                            cell.alignment = Alignment(horizontal='left', vertical='center')
                        
                        cell.border = border
                
                # Adjust column widths
                for col_idx, column in enumerate(detail_df.columns, 1):
                    max_length = min(max(
                        detail_df[column].astype(str).map(len).max(),
                        len(str(column))
                    ) + 2, 35)
                    column_letter = ws2.cell(row=1, column=col_idx).column_letter
                    ws2.column_dimensions[column_letter].width = max_length
                
                print(f" PR Approval details rows: {len(detail_df)}")
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"{selected_division}_PrApproval_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        print(f" PR Approval export completed: {filename}")
        
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        print(f" PR Approval export error: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Export error: {str(e)}")

@app.get("/api/captcha")
async def get_captcha():
    """Generate and return a CAPTCHA"""
    captcha_text, captcha_image = CaptchaGenerator.generate_captcha()
    return {
        "captcha": captcha_text,
        "image": captcha_image
    }

@app.get("/api/vehicle-images")
async def get_vehicle_images():
    """Return vehicle images"""
    images_data = [{'name': img['name'], 'data': img['data']} for img in MAHINDRA_IMAGES]
    return {"images": images_data}

@app.post("/api/login")
async def api_login(request: Request):
    """Handle API login request"""
    try:
        body = await request.json()
        user_id = body.get('user_id', '').strip()
        password = body.get('password', '')
        
        if not user_id or user_id not in USER_CREDENTIALS:
            raise HTTPException(status_code=401, detail="Invalid User ID")
        
        if USER_CREDENTIALS[user_id] != password:
            raise HTTPException(status_code=401, detail="Invalid Password")
        
        session_id = create_session(user_id)
        
        response_data = {
            "success": True,
            "session_id": session_id,
            "user_id": user_id,
            "message": "Login successful"
        }
        
        response = JSONResponse(content=response_data, status_code=200)
        response.set_cookie(
            key="session_id", 
            value=session_id, 
            httponly=True, 
            max_age=28800,
            samesite="lax",
            path="/"
        )
        return response
        
    except HTTPException as e:
        raise
    except Exception as e:
        print(f"Login error: {e}")
        raise HTTPException(status_code=400, detail=str(e))

@app.get("/api/warranty-data")
async def get_warranty_data():
    """Get warranty data (Credit, Debit, Arbitration, Current Month)"""
    try:
        print(f" Warranty data request received")
        
        if WARRANTY_DATA['credit_df'] is None:
            print(f" Warranty data not loaded")
            return {
                "credit": [],
                "debit": [],
                "arbitration": [],
                "currentMonth": [],
                "compensation": [],
                "prApproval": []
            }
        
        print(f" Processing warranty data...")
        credit_records = WARRANTY_DATA['credit_df'].to_dict('records')
        debit_records = WARRANTY_DATA['debit_df'].to_dict('records')
        arbitration_records = WARRANTY_DATA['arbitration_df'].to_dict('records')
        
        # Process current month warranty data
        current_month_records = []
        if WARRANTY_DATA['current_month_df'] is not None:
            current_month_records = WARRANTY_DATA['current_month_df'].to_dict('records')
        
        # Process compensation claim data
        compensation_records = []
        if WARRANTY_DATA['compensation_df'] is not None:
            compensation_records = WARRANTY_DATA['compensation_df'].to_dict('records')
        
        # Process PR Approval data
        pr_approval_records = []
        if WARRANTY_DATA['pr_approval_df'] is not None:
            pr_approval_records = WARRANTY_DATA['pr_approval_df'].to_dict('records')
        
        for records in [credit_records, debit_records, arbitration_records, current_month_records, compensation_records, pr_approval_records]:
            for record in records:
                for key in record:
                    if pd.isna(record[key]):
                        record[key] = 0
        
        print(f"   Warranty data prepared successfully")
        print(f"   Credit rows: {len(credit_records)}")
        print(f"   Debit rows: {len(debit_records)}")
        print(f"   Arbitration rows: {len(arbitration_records)}")
        print(f"   Current Month rows: {len(current_month_records)}")
        print(f"   Compensation rows: {len(compensation_records)}")
        print(f"   PR Approval rows: {len(pr_approval_records)}")
        
        return {
            "credit": credit_records,
            "debit": debit_records,
            "arbitration": arbitration_records,
            "currentMonth": current_month_records,
            "compensation": compensation_records,
            "prApproval": pr_approval_records
        }
    except Exception as e:
        print(f" Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/login-page")
async def login_page():
    """Serve the login page"""
    return HTMLResponse(content=LOGIN_PAGE)

@app.get("/dashboard")
async def dashboard():
    """Serve dashboard (no login required)"""
    return HTMLResponse(content=DASHBOARD_HTML)


@app.get("/")
async def root():
    """Root route - directly serve dashboard (no login required)"""
    return HTMLResponse(content=DASHBOARD_HTML)

# ==================== STARTUP ====================

print("\n" + "=" * 100)
print("STARTING WARRANTY MANAGEMENT SYSTEM - PORT 8001")
print("=" * 100)

print("\nProcessing warranty data...")
WARRANTY_DATA['credit_df'], WARRANTY_DATA['debit_df'], WARRANTY_DATA['arbitration_df'], WARRANTY_DATA['source_df'] = process_warranty_data()

print("\nProcessing current month warranty data...")
WARRANTY_DATA['current_month_df'], WARRANTY_DATA['current_month_source_df'] = process_current_month_warranty()

print("\nProcessing compensation claim data...")
WARRANTY_DATA['compensation_df'], WARRANTY_DATA['compensation_source_df'] = process_compensation_claim()

print("\nProcessing PR Approval data...")
WARRANTY_DATA['pr_approval_df'], WARRANTY_DATA['pr_approval_source_df'] = process_pr_approval()

if __name__ == "__main__":
    hostname = socket.gethostname()
    try:
        local_ip = socket.gethostbyname(hostname)
    except:
        local_ip = "127.0.0.1"
    
    port = 8001
    
    print("\n" + "=" * 100)
    print(f" SERVER READY - Warranty Dashboard")
    print("=" * 100)
    print(f" PORT: 8001")
    print(f" Login URL: http://localhost:{port}/login-page")
    print(f" Network URL: http://{local_ip}:{port}/login-page")
    print(f"\n Test Credentials:")
    print(f"   User ID: 11724")
    print(f"   Password: un001@123")
    print("\n" + "=" * 100 + "\n")
    
    uvicorn.run(app, host="0.0.0.0", port=port)
